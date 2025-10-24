from django.shortcuts import render, redirect, get_object_or_404
from django.core.files.storage import FileSystemStorage
from .models import *
import pandas as pd
from .forms import PreEnteredDOBulkForm
from django.http import HttpResponse,JsonResponse
from django.conf import settings
import requests
from django.core.paginator import Paginator
from django.urls import reverse
from django.contrib import messages
# from twilio.rest import Client
from django.contrib.auth.views import LoginView
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
import pytz
from django.utils import timezone
from django.http import HttpResponseForbidden
from functools import wraps
from django.db import IntegrityError
from django.contrib.auth import login,authenticate
from django.db.models import Q
from datetime import datetime
from datetime import timedelta




import openpyxl

# Bulk update DO status view
from django.contrib.auth.decorators import login_required
@login_required
def bulk_update_do_status(request):
    messages_list = []
    status_choices = DeliveryOrder.DO_STATUS_CHOICES
    if request.method == 'POST':
        status = request.POST.get('status')
        excel_file = request.FILES.get('excel_file')
        if not excel_file or not status:
            messages_list.append('Please upload an Excel file and select a status.')
        else:
            try:
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                do_numbers = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    do_number = str(row[0]).strip() if row[0] else None
                    if do_number:
                        do_numbers.append(do_number)
                updated = DeliveryOrder.objects.filter(do_number__in=do_numbers).update(status=status)
                messages_list.append(f"✅ Updated {updated} delivery orders to status '{status}'.")
            except Exception as e:
                messages_list.append(f"❌ Error processing file: {e}")
    return render(request, 'orders/bulk_update_do_status.html', {
        'status_choices': status_choices,
        'messages': messages_list
    })

def custom_404(request, exception):
    return render(request, 'credit/404.html', status=404)


def home(request):
    return render(request, 'orders/home.html')

def role_required(*required_roles):  # Accept multiple roles
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return HttpResponseForbidden("You are not authorized to view this page.")

            try:
                user_role = request.user.role.role  # Assuming `role` is a related object
                if user_role in required_roles:  # Check if user role is in allowed roles
                    return view_func(request, *args, **kwargs)
            except AttributeError:  # If role doesn't exist
                pass

            return HttpResponseForbidden("You are not authorized to view this page.")
        return _wrapped_view
    return decorator


@login_required
def dashboard(request):
    return render(request, 'orders/dashboard.html')

class CustomLoginView(LoginView):
    template_name = 'orders/login.html'
    redirect_authenticated_user = True

    def form_invalid(self, form):
        messages.error(self.request, "❌ Invalid username or password!")
        return self.render_to_response(self.get_context_data(form=form))

    def get_success_url(self):
        try:
            user_role = self.request.user.role.role
        except AttributeError:
            return reverse_lazy('home')  # Default redirect

        role_redirects = {
            'Junaid Admin':'order_list',
            'Admin': 'order_list',
            'Warehouse': 'vehicle_list',
            'Security': 'security_vehicle_list',
            'Salesman': 'salesman_orders',
            'Driver': 'driver_vehicle_list',
            'Accounts': 'account_delivered_orders',
            'Collection': 'customer_list',
            'Manager': 'md_dashboard',
        }
        return reverse_lazy(role_redirects.get(user_role, 'home'))

from django.contrib.auth.views import LogoutView

class CustomLogoutView(LogoutView):
    next_page = 'home'


def upload_file(request):
    messages = []  # Store error messages

    if request.method == 'POST':
        # Check if the file is uploaded
        if 'excel_file' not in request.FILES:
            messages.append("❌ Please upload the Excel file!")
            return render(request, 'orders/upload_file.html', {'messages': messages})

        fs = FileSystemStorage()

        try:
            excel_file = request.FILES['excel_file']
            filename = fs.save(excel_file.name, excel_file)
            file_path = fs.path(filename)

            # Use a context manager to open the Excel file
            with pd.ExcelFile(file_path, engine='openpyxl') as xls:
                # Process Delivery Orders Sheet
                try:
                    df = pd.read_excel(xls, sheet_name='DO')
                    df['DO'] = df['DO'].apply(lambda x: str(int(x)) if isinstance(x, float) and x.is_integer() else str(x))
                    df['INVOICE'] = df['INVOICE'].apply(lambda x: str(int(x)) if isinstance(x, float) and x.is_integer() else str(x))
                    df['Delivery Mobile'] = df['Delivery Mobile'].apply(lambda x: str(int(x)) if isinstance(x, float) and x.is_integer() else str(x))
                    df['Sales Person Mobile'] = df['Sales Person Mobile'].apply(lambda x: str(int(x)) if isinstance(x, float) and x.is_integer() else str(x))
                    df['INVOICE'] = df['INVOICE'].apply(lambda x: f"C{uuid.uuid4().hex[:12].upper()}" if str(x).strip().upper() == 'NIL' else x)
                    for index, row in df.iterrows():
                        try:
                            DeliveryOrder.objects.create(
                                do_number=row['DO'],
                                date=row['DATE'],
                                customer_code=row['CUSTOMER CODE'],
                                customer_name = row['Debit Customer'] if row['CUSTOMER'] == "DEBIT CUSTOMER ( CASH )" else row['CUSTOMER'],
                                salesman=row['Salesman'],
                                mobile_number=row['Delivery Mobile'],
                                salesman_mobile=row['Sales Person Mobile'],
                                city=row['CITY'],
                                area=row['AREA'],
                                lpo=row['LPO'],
                                invoice_number=row['INVOICE'],
                                amount=row['AMOUNT'],
                            )
                        except IntegrityError:
                            messages.append(f"❌ Duplicate DO Number: {row['DO']} already exists! Others are uploaded")
                except Exception as e:
                    messages.append(f"❌ Error processing 'DO' sheet: {str(e)}")

                # Process Item Data Sheet
                try:
                    df_items = pd.read_excel(xls, sheet_name='DO_Itemwise')

                    # Convert DO Number to string and remove trailing .0
                    df_items['Document Number'] = df_items['Document Number'].apply(lambda x: str(int(x)) if isinstance(x, float) and x.is_integer() else str(x))

                    for index, row in df_items.iterrows():
                        try:
                        # Check if the item already exists for the DO
                            existing_item = DeliveryItemWise.objects.filter(
                            do_number=row['Document Number'],
                            item_code=row['Item No.']
                            ).exists()

                            if not existing_item:
                                DeliveryItemWise.objects.create(
                                do_number=row['Document Number'],
                                item_code=row['Item No.'],
                                item_description=row['Item/Service Description'],
                                quantity=row['Quantity'],
                                price=row['Price'],
                            )
                            else:
                                messages.append(f"❌ Item {row['Item No.']} for DO {row['Document Number']} already exists!")
                        except IntegrityError:
                            messages.append(f"❌ Integrity Error: Duplicate Item for DO: {row['Document Number']} and Item Code: {row['Item No.']} already exists!")
                except Exception as e:
                    messages.append(f"❌ Error processing 'DO_Itemwise' sheet: {str(e)}")
            pending_pre_dos = PreEnteredDO.objects.filter(delivered=False)
            for pre_do in pending_pre_dos:
                matched_do = DeliveryOrder.objects.filter(do_number=pre_do.do_number).first()
                if matched_do:
                    matched_do.status = 'Delivered'
                    matched_do.save()
                    pre_do.delivered = True
                    pre_do.save()

            fs.delete(filename)  # Delete file after processing

        except Exception as e:
            messages.append(f"❌ Error processing Excel file: {str(e)}")

        # If no errors, redirect
        if not messages:
            return redirect('order_list')

    return render(request, 'orders/upload_file.html', {'messages': messages})

@login_required
@role_required('Security')
def security_vehicle_list(request):
    vehicles = Vehicle.objects.exclude(id__in=[12, 13,22,21])
    return render(request, 'orders/security_vehicle_list.html', {'vehicles': vehicles})


@login_required
@role_required('Security')
def update_vehicle_status(request, vehicle_id):
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)

    # Check if any orders for this vehicle have the status 'Loaded'
    orders = DeliveryOrder.objects.filter(vehicle=vehicle, status='Loaded')

    if request.method == 'POST':
        new_status = request.POST.get('status')

        # Update only the orders where the current status is 'Loaded'
        if orders.exists():
            orders.update(status=new_status, delivery_date=None)
            # Optionally, handle specific logic for sending messages, etc.
            send_out_for_delivery_messages(vehicle)

        return redirect('security_vehicle_list')

    return render(request, 'orders/update_vehicle_status.html', {'vehicle': vehicle})

@login_required
@role_required('Warehouse')
def vehicle_list(request):
    vehicles = Vehicle.objects.filter(Q(id__in=[1, 2, 3, 4, 17,23]))
    return render(request, 'orders/vehicle_list.html', {'vehicles': vehicles})
from django.db import models  # Import models for Q objects
from django.db.models import Q  # Import Q for complex queries

@login_required
@role_required('Admin', 'Junaid Admin')
def order_list(request):
    # Check if the 'hide_delivered' parameter is present in the request
    hide_delivered = request.GET.get('hide_delivered', 'false').lower() == 'true'

    # Start with all orders
    orders = DeliveryOrder.objects.all().order_by('-date')

    # Filter by hide_delivered
    # if hide_delivered:
    #     orders = orders.exclude(status='Delivered')

    # Filter by date if provided
    # Filter by date if provided
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if from_date and to_date:
        orders = orders.filter(date__range=[from_date, to_date])
    elif from_date:
        orders = orders.filter(date__gte=from_date)
    elif to_date:
        orders = orders.filter(date__lte=to_date)
    date_filter = request.GET.get('date')
    if date_filter:
        orders = orders.filter(date=date_filter)

    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter == "delivered_group":
        orders = orders.filter(status__in=["Delivered", "Received by A/c"])
    elif status_filter:
        orders = orders.filter(status=status_filter)

    driver_filter = request.GET.get('driver')
    if driver_filter:
        orders = orders.filter(driver=driver_filter)

    salesman = request.GET.get('salesman')
    if salesman:
        orders = orders.filter(salesman=salesman)

    city = request.GET.get('city')
    if city:
        orders = orders.filter(city=city)

    # Apply search query
    search_query = request.GET.get('search_query')
    if search_query:
        # Filter by DO Number, Customer Name, or Mobile
        orders = orders.filter(
            models.Q(do_number__icontains=search_query) |
            models.Q(customer_name__icontains=search_query) |
            models.Q(mobile_number__icontains=search_query)  |
            models.Q(invoice_number__icontains=search_query)
        )


    salesmen = DeliveryOrder.objects.values_list('salesman', flat=True).distinct()
    citys = DeliveryOrder.objects.values_list('city', flat=True).distinct()

    vehicle_filter = request.GET.get('vehicle')
    if vehicle_filter:
        orders = orders.filter(vehicle_id=vehicle_filter)
    # Pagination
    paginator = Paginator(orders, 300)  # Show 150 orders per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    docount=DeliveryOrder.objects.all().count()
    pending_count = orders.filter(status='Pending').count()
    delivered_count = orders.filter(status__in=['Delivered','Received by A/c']).count()
    out_for_delivery_count=orders.filter(status="Out for Delivery").count()
    onhold_count= orders.filter(status="On Hold").count()

    vehicles = Vehicle.objects.all()
    return render(request, 'orders/order_list.html', {
        'orders': page_obj,
        'vehicles': vehicles,
        'hide_delivered': hide_delivered,
        'docount': docount,
        'pending_count': pending_count,
        'delivered_count': delivered_count,
        'out_for_delivery_count': out_for_delivery_count,
        'onhold_count':onhold_count,
        'salesmen': salesmen,
        'citys': citys,
        'selected_salesman': salesman,
        'selected_city': city,
        'selected_vehicle': vehicle_filter,
        'search_query': search_query,

    })

@login_required
@role_required('Admin')
def update_order(request, do_number):
    # Use get_object_or_404 to fetch the order by do_number
    order = get_object_or_404(DeliveryOrder, do_number=do_number)
    if request.method == 'POST':
        vehicle_id = request.POST.get('vehicle')
        new_status = request.POST.get('status')
        mobile_number = request.POST.get('mobile_number')
        driver=request.POST.get('driver')
        invoice_number= request.POST.get('invoice_number')
        amount= request.POST.get('amount')
        order.vehicle_id = vehicle_id
        order.mobile_number = mobile_number
        order.status = new_status
        order.driver= driver
        order.invoice_number = invoice_number or None
        order.amount= amount


        if new_status == 'Delivered' or new_status == 'Partial Delivery':
            order.delivery_date = timezone.now().date()
            print(f"✅ Order {do_number} marked as Delivered by {request.user.username}")
        else:
            print(f"⚠️ Order {do_number} changed from {order.status} to {new_status} by {request.user.username}")
            order.delivery_date = None
        order.save()

        # Check if the status is changing to "Delivered"
        # if new_status == 'Delivered':
        #     send_whatsapp_message(order.mobile_number, order.do_number)

        return redirect('order_list')
    vehicles = Vehicle.objects.all()
    return render(request, 'orders/update_order.html', {'order': order, 'vehicles': vehicles})
@login_required
@role_required('Warehouse','Admin')
def update_vehicle(request, vehicle_id):
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)

    if request.method == 'POST':
        # Get the selected DO numbers as a comma-separated string and split it into a list
        do_numbers = request.POST.get('do_numbers', '').split(',')

        # Get the selected driver
        driver = request.POST.get('driver')
        new_driver = 'Customer' if vehicle and vehicle_id == 12 else driver

        new_status = 'Delivered' if vehicle and vehicle.id == 12 else 'Partial Delivery' if vehicle and vehicle.vehicle_number=='SELF PARTIAL' else 'On Hold' if vehicle and vehicle.vehicle_number == 'HOLD' else 'GRV' if vehicle and vehicle.id == 13 else 'Cancelled' if vehicle and vehicle.vehicle_number=='Cancelled' else 'Loaded'
        uae_tz = pytz.timezone('Asia/Dubai')
        new_date = timezone.now().astimezone(uae_tz) if vehicle and vehicle.id == 12 else None

        DeliveryOrder.objects.filter(
            vehicle=vehicle
        ).exclude(  # Exclude delivered orders
            status__in=['Delivered', 'Partial Delivery', 'Not Delivered', 'Cancelled','GRV','Received by A/c','On Hold']
        ).update(
            vehicle=None,
            status='Pending',
            driver=None
        )

        # Update the selected DOs
        orders = DeliveryOrder.objects.filter(do_number__in=do_numbers)
        for order in orders:
            order.vehicle = vehicle
            order.status = new_status
            order.driver = new_driver
            order.delivery_date = new_date if new_status in ['Delivered', 'Partial Delivery'] else None
            order.save()

        return redirect('vehicle_list')
    pending_count = DeliveryOrder.objects.filter(status='Pending').count()
    not_delivered_count = DeliveryOrder.objects.filter(status='Not Delivered').count()
    partial_count=DeliveryOrder.objects.filter(status='Partial Delivery').count()
    # Get all DOs that are not yet delivered and not assigned to any vehicle
    available_orders = DeliveryOrder.objects.filter(
        status__in=['Pending', 'Partial Delivery', 'Not Delivered','On Hold'],
        delivery_date=None
    )

    # Get DOs already assigned to this vehicle
    assigned_orders = DeliveryOrder.objects.filter(vehicle=vehicle).exclude(status__in=['Delivered','On Hold','Cancelled','Partial Delivery', 'Not Delivered','Received by A/c','GRV'])

    return render(request, 'orders/update_vehicle.html', {
        'vehicle': vehicle,
        'pending_count':pending_count,
        'partial_count':partial_count,
        'not_delivered_count':not_delivered_count,
        'available_orders': available_orders,
        'assigned_orders': assigned_orders,
        'drivers': DeliveryOrder.DRIVER_CHOICES
    })

@login_required
@role_required('Admin', 'Junaid Admin')
def export_orders_to_excel(request):
    orders = DeliveryOrder.objects.all().order_by('-date')

    # Apply filters like in order_list
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if from_date and to_date:
        orders = orders.filter(date__range=[from_date, to_date])
    elif from_date:
        orders = orders.filter(date__gte=from_date)
    elif to_date:
        orders = orders.filter(date__lte=to_date)

    date_filter = request.GET.get('date')
    if date_filter:
        orders = orders.filter(date=date_filter)

    status_filter = request.GET.get('status')
    if status_filter == "delivered_group":
        orders = orders.filter(status__in=["Delivered", "Received by A/c"])
    elif status_filter:
        orders = orders.filter(status=status_filter)

    driver_filter = request.GET.get('driver')
    if driver_filter:
        orders = orders.filter(driver=driver_filter)

    salesman = request.GET.get('salesman')
    if salesman:
        orders = orders.filter(salesman=salesman)

    city = request.GET.get('city')
    if city:
        orders = orders.filter(city=city)

    vehicle_filter = request.GET.get('vehicle')
    if vehicle_filter:
        orders = orders.filter(vehicle_id=vehicle_filter)

    search_query = request.GET.get('search_query')
    if search_query:
        orders = orders.filter(
            models.Q(do_number__icontains=search_query) |
            models.Q(customer_name__icontains=search_query) |
            models.Q(mobile_number__icontains=search_query)
        )

    # Build DataFrame
    data = {
        'DO': [order.do_number for order in orders],
        'DATE': [order.date for order in orders],
        'CUSTOMER CODE': [order.customer_code for order in orders],
        'CUSTOMER': [order.customer_name for order in orders],
        'MOBILE': [order.mobile_number if order.mobile_number else '' for order in orders],
        'SALESPERSON': [order.salesman if order.salesman else '' for order in orders],
        'VEHICLE': [order.vehicle.vehicle_number if order.vehicle else '' for order in orders],
        'INVOICE': [order.invoice_number if order.invoice_number else '' for order in orders],
        'AMOUNT':[order.amount if order.amount else '' for order in orders],
        'DRIVER' : [order.driver if order.driver else '' for order in orders],
        'STATUS': [order.status for order in orders],
        'AREA': [order.area for order in orders]
    }
    df = pd.DataFrame(data)

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=delivery_orders_filtered.xlsx'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Orders')

    return response
@login_required
@role_required('Admin')
def delete_order(request, do_number):
    # Use get_object_or_404 to retrieve the order or return a 404 if not found
    order = get_object_or_404(DeliveryOrder, do_number=do_number)

    # Delete the order
    order.delete()

    # Redirect to the order list page after deletion
    return redirect(reverse('order_list'))

@login_required
@role_required('Admin')
def delete_all_orders(request):
    # Delete all DeliveryOrder records
    DeliveryOrder.objects.all().delete()
    DeliveryItemWise.objects.all().delete()

    # Optionally, add a success message
    # messages.success(request, "All delivery orders have been deleted.")

    # Redirect to the order list page after deletion
    return redirect(reverse('order_list'))

# @login_required
# # @role_required('Admin')
# def send_whatsapp_message(to_number, do_number):
#     # Twilio credentials
#     account_sid = settings.TWILIO_ACCOUNT_SID
#     auth_token = settings.TWILIO_AUTH_TOKEN
#     client = Client(account_sid, auth_token)

#     # Format the phone number for WhatsApp
#     whatsapp_number = f'whatsapp:{to_number}'

#     # Send the message
#     message = client.messages.create(
#         body=f"Your delivery order {do_number} has been delivered. Thank you for choosing us!",
#         from_=f'whatsapp:{settings.TWILIO_WHATSAPP_NUMBER}',
#         to=whatsapp_number
#     )

@login_required
@role_required('Salesman')
def salesman_orders(request):
    # Get the logged-in user's username
    username = request.user.username
    firstname = request.user.first_name

    # Get all orders for this salesman using case-insensitive filtering
    orders = DeliveryOrder.objects.filter(salesman__iexact=firstname).order_by('-date')

    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter:
        orders = orders.filter(status=status_filter)

    # Pagination
    paginator = Paginator(orders, 10)  # Show 15 orders per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'orders/salesman_orders.html', {
        'page_obj': page_obj,
        'status_filter': status_filter,
    })

@login_required
@role_required('Driver')
def driver_vehicle_list(request):
    # List all vehicles
    vehicles = Vehicle.objects.exclude(id=17)
    return render(request, 'orders/driver_vehicle_list.html', {
        'vehicles': vehicles,
    })

@login_required
@role_required('Driver')
def driver_vehicle_list(request):
    # List all vehicles
    vehicles = Vehicle.objects.exclude(vehicle_number__in=['GRV/Cancel', 'GRV / Cancel','Self Pickup'])
    return render(request, 'orders/driver_vehicle_list.html', {
        'vehicles': vehicles,
    })


@login_required
@role_required('Driver')
def update_do_status(request, vehicle_id):
    # Get the selected vehicle
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)

    # Get all DOs assigned to this vehicle with status 'Out for Delivery'
    orders = DeliveryOrder.objects.filter(vehicle=vehicle, status='Out for Delivery')

    if request.method == 'POST':
        # Update the status of a single DO
        do_number = request.POST.get('do_number')
        new_status = request.POST.get('status')
        order = get_object_or_404(DeliveryOrder, do_number=do_number, vehicle=vehicle)
        order.status = new_status
        if new_status == 'Delivered':
            uae_tz = pytz.timezone('Asia/Dubai')
            order.delivery_date = timezone.now().astimezone(uae_tz)
            # send_whatsapp_message(order.salesman_mobile, order.do_number)
        else:
            order.delivery_date = None
        order.save()
        return redirect('update_do_status', vehicle_id=vehicle.id)

    return render(request, 'orders/update_do_status.html', {
        'vehicle': vehicle,
        'orders': orders,
    })

@login_required
@role_required('Warehouse')
def pending_do_list(request):
    # Get all delivery orders with statuses 'Pending', 'Partial Delivery', and any other non-delivered status
    pending_orders = DeliveryOrder.objects.filter(status__in=['Pending', 'Partial Delivery', 'Not Delivered'])

    return render(request, 'orders/pending_do_list.html', {
        'pending_orders': pending_orders,
    })

@login_required
@role_required('Security')
def security_verify(request, vehicle_id):
    # Get all delivery orders with statuses 'Pending', 'Partial Delivery', and any other non-delivered status
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)
    security_orders = DeliveryOrder.objects.filter(vehicle=vehicle,status='Loaded')

    return render(request, 'orders/security_verify.html', {
        'security_orders': security_orders,
    })


@login_required
# @role_required('Salesman')
def do_items(request, do_number):
    # Get all items for the specified DO number
    items = DeliveryItemWise.objects.filter(do_number=do_number)
    print(f"Items for DO {do_number}: {items}")  # Debugging line

    return render(request, 'orders/do_items.html', {
        'items': items,
        'do_number': do_number,
    })


####################################       WHATSAPP        #####################################'
import requests
from django.conf import settings
def send_out_for_delivery_messages(vehicle):
    print(f"Sending Out for Delivery Messages for Vehicle: {vehicle}")

    api_url = f"https://graph.facebook.com/v22.0/{settings.WHATSAPP_PHONE_NUMBER_ID}/messages"
    access_token = settings.WHATSAPP_ACCESS_TOKEN

    orders = DeliveryOrder.objects.filter(vehicle=vehicle, status='Out for Delivery')
    print(f"Orders Found: {orders.count()}")  # New Print

    if not orders.exists():
        print("No orders with 'Out for Delivery' status for this vehicle.")
        return

    for order in orders:
        print(f"Preparing message for Order: {order.do_number}")

        whatsapp_number = f"{order.mobile_number}"
        if not whatsapp_number.startswith('+'):
            whatsapp_number = f"+{whatsapp_number}"  # Auto-fix missing '+'

        do_number = order.do_number

        payload = {
    "messaging_product": "whatsapp",
    "to": whatsapp_number,  # Recipient's phone number
    "type": "template",
    "template": {
        "name": "out_for_delivery",  # Template name
        "language": {
            "code": "en"  # Language code
        },
        "components": [
            {
                "type": "BODY",
                "parameters": [
                    {
                        "type": "text",
                        "parameter_name": "do_number",
                        "text": str(do_number)  # Ensure do_number is passed as a string
                    }
                ]
            }
        ]
    }
}

        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json"
        }

        # Send the POST request to the API
        response = requests.post(api_url, headers=headers, json=payload)
        print(response.status_code)
        print(response.json())  # This will give you the complete response including any status

        # Check and log the response
        if response.status_code == 200:
            print(f"Message sent to {whatsapp_number} successfully!")
        else:
            print(f"Failed to send message to {whatsapp_number}: {response.text}")


# def send_out_for_delivery_messages(vehicle):
#     # WhatsApp API URL for sending messages
#     api_url = f"https://graph.facebook.com/v22.0/{settings.WHATSAPP_PHONE_NUMBER_ID}/messages"

#     # API Token from Meta (WhatsApp Cloud API)
#     access_token = settings.WHATSAPP_ACCESS_TOKEN

#     # Get all delivery orders for the vehicle
#     orders = DeliveryOrder.objects.filter(vehicle=vehicle,status='Out for Delivery')

#     for order in orders:
#         # Format the phone number for WhatsApp (ensure the mobile number includes the country code, e.g., +971 for UAE)
#         whatsapp_number = f"{order.mobile_number}"  # Replace `order.salesman_mobile` with the correct field
#         do_number = order.do_number  # Ensure this is the correct field for order number
#         lpo_number = order.lpo  # Make sure this field exists on your model

#         # Define the message payload using a template with named parameters
#         payload = {
#     "messaging_product": "whatsapp",
#     "to": whatsapp_number,  # Recipient's phone number
#     "type": "template",
#     "template": {
#         "name": "out_for_delivery",  # Template name
#         "language": {
#             "code": "en"  # Language code
#         },
#         "components": [
#             {
#                 "type": "BODY",
#                 "parameters": [
#                         {
#                             "type": "text",
#                             "text": str(do_number)
#                         }
#                     ]
#             }
#         ]
#     }
# }

#         # Headers for authorization
#         headers = {
#             "Authorization": f"Bearer {access_token}",
#             "Content-Type": "application/json"
#         }

#         # Send the POST request to the API
#         response = requests.post(api_url, headers=headers, json=payload)
#         print(response.status_code)
#         print(response.json())  # This will give you the complete response including any status

#         # Check and log the response
#         if response.status_code == 200:
#             print(f"Message sent to {whatsapp_number} successfully!")
#         else:
#             print(f"Failed to send message to {whatsapp_number}: {response.text}")

@login_required
# @role_required('Admin')
def add_vehicle(request):
    # Handle adding a new vehicle
    if request.method == 'POST' and 'add_vehicle' in request.POST:
        vehicle_number = request.POST.get('vehicle_number')
        Vehicle.objects.create(vehicle_number=vehicle_number)
        return redirect('add_vehicle')  # Redirect to the same page to show updated list

    # Handle deleting a vehicle
    if request.method == 'POST' and 'delete_vehicle' in request.POST:
        vehicle_id = request.POST.get('vehicle_id')
        vehicle = get_object_or_404(Vehicle, id=vehicle_id)
        vehicle.delete()
        return redirect('add_vehicle')  # Redirect to show updated list after deletion

    # Get all vehicles to display in the template
    vehicles = Vehicle.objects.exclude(id__in=[12, 13,17,22,21])[6:]

    return render(request, 'orders/add_vehicle.html', {'vehicles': vehicles})



from django.http import JsonResponse
from django.db.models import Q

@login_required
@role_required('Admin', 'Junaid Admin')
def order_search(request):
    # Get the search term from the request
    search_term = request.GET.get('search', '').strip().lower()

    # Filter orders based on the search term
    orders = DeliveryOrder.objects.filter(
        Q(do_number__icontains=search_term) |
        Q(customer_name__icontains=search_term) |
        Q(mobile_number__icontains=search_term) |
        Q(salesman__icontains=search_term) |
        Q(city__icontains=search_term)
    ).order_by('-date')

    # Serialize the filtered orders into JSON
    data = []
    for order in orders:
        data.append({
            'do_number': order.do_number,
            'date': order.date.strftime('%Y-%m-%d'),
            'customer_name': order.customer_name,
            'mobile_number': order.mobile_number,
            'salesman': order.salesman,
            'salesman_mobile': order.salesman_mobile,
            'city': order.city,
            'area': order.area,
            'vehicle': str(order.vehicle),
            'driver': order.driver,
            'status': order.status,
            'delivery_date': order.delivery_date.strftime('%Y-%m-%d') if order.delivery_date else None,
        })

    return JsonResponse(data, safe=False)



import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

import logging
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

# Set up a logger
logger = logging.getLogger(__name__)

@csrf_exempt
def whatsapp_webhook(request):
    if request.method == "GET":
        verify_token = "junaid_whatsapp_webhook"
        challenge = request.GET.get("hub.challenge")
        token = request.GET.get("hub.verify_token")
        if token == verify_token:
            return JsonResponse(int(challenge), safe=False)
        return JsonResponse({"error": "Invalid verification token"}, status=403)

    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            logger.info(f"Received webhook data: {json.dumps(data, indent=4)}")

            if "entry" in data:
                for entry in data["entry"]:
                    for change in entry.get("changes", []):
                        if change["field"] == "messages":
                            message_data = change["value"]

                            # 📩 Process customer replies (incoming messages)
                            if "messages" in message_data:
                                for message in message_data["messages"]:
                                    message_id = message.get("id", "UNKNOWN_ID")
                                    sender = message.get("from", "UNKNOWN_SENDER")

                                    if message.get("type") == "text":
                                        text_body = message["text"].get("body", "")

                                        # Save to database
                                        CustomerReply.objects.create(
                                            sender=sender,
                                            message_id=message_id,
                                            text_body=text_body
                                        )
                                        logger.info(f"Stored customer reply from {sender}: {text_body}")

                                        reply_message = reply_message = (
    "*💬 Thank you for reaching out to Junaid Group!* \n\n"
    "📞 *For any inquiries, please contact us directly or WhatsApp us at +971 558946349*.\n\n"
    " This is an autogenerated message. Please do not reply to this chat."
)
                                        send_whatsapp_message(sender, reply_message)

                            # ✅ Process sent message statuses
                            if "statuses" in message_data:
                                for status in message_data["statuses"]:
                                    message_id = status.get("id", "UNKNOWN_ID")
                                    status_text = status.get("status", "unknown")
                                    timestamp = status.get("timestamp", "")
                                    recipient_id = status.get("recipient_id", "UNKNOWN_RECIPIENT")

                                    # 🛠️ Check if message_id already exists before inserting
                                    try:
                                        MessageStatus.objects.create(
                                        message_id=message_id,
                                        recipient_id=recipient_id,
                                        status=status_text,
                                        timestamp=timestamp
                                        )
                                        logger.info(f"Message {message_id} to {recipient_id} is {status_text}")

                                    except IntegrityError:
                                        logger.warning(f"Duplicate message_id detected: {message_id}, skipping insertion.")

            return JsonResponse({"success": True}, status=200)

        except json.JSONDecodeError:
            logger.error("Invalid JSON in the request body")
            return JsonResponse({"error": "Invalid JSON"}, status=400)

api_url = f"https://graph.facebook.com/v22.0/{settings.WHATSAPP_PHONE_NUMBER_ID}/messages"

def send_whatsapp_message(recipient_id, message_text):
    headers = {
        "Authorization": f"Bearer {settings.WHATSAPP_ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }
    data = {
        "messaging_product": "whatsapp",
        "to": recipient_id,
        "type": "text",
        "text": {"body": message_text}
    }
    response = requests.post(api_url, headers=headers, json=data)
    logger.info(f"Sent WhatsApp message to {recipient_id}: {message_text}, Response: {response.text}")

# @csrf_exempt
# def whatsapp_webhook(request):
#     if request.method == "GET":
#         # WhatsApp requires verification when setting up a webhook
#         verify_token = "junaid_whatsapp_webhook"  # Choose any random token and use it in Meta settings
#         challenge = request.GET.get("hub.challenge")
#         token = request.GET.get("hub.verify_token")

#         if token == verify_token:
#             return JsonResponse(int(challenge), safe=False)
#         return JsonResponse({"error": "Invalid verification token"}, status=403)

#     elif request.method == "POST":
#         try:
#             data = json.loads(request.body)

#             # Log the incoming data for inspection
#             logger.info(f"Received webhook data: {json.dumps(data, indent=4)}")

#             # Check if the structure is valid and contains the necessary data
#             if "messages" in data:
#                 for message in data["messages"]:
#                     # Extract message status
#                     message_status = message.get("message_status")
#                     if message_status:
#                         logger.info(f"Message status: {message_status}")
#                         # You can also store this in the database or perform other actions
#                     else:
#                         logger.error("Message status is missing")

#             return JsonResponse({"success": True}, status=200)

#         except json.JSONDecodeError:
#             logger.error("Invalid JSON in the request body")
#             return JsonResponse({"error": "Invalid JSON"}, status=400)

# @csrf_exempt
# def whatsapp_webhook(request):
#     if request.method == "GET":
#         verify_token = "junaid_whatsapp_webhook"
#         challenge = request.GET.get("hub.challenge")
#         token = request.GET.get("hub.verify_token")
#         if token == verify_token:
#             return JsonResponse(int(challenge), safe=False)
#         return JsonResponse({"error": "Invalid verification token"}, status=403)

#     elif request.method == "POST":
#         try:
#             data = json.loads(request.body)
#             logger.info(f"Received webhook data: {json.dumps(data, indent=4)}")

#             if "entry" in data:
#                 for entry in data["entry"]:
#                     for change in entry.get("changes", []):
#                         if change["field"] == "messages":
#                             message_data = change["value"]

#                             # Process customer replies
#                             if "messages" in message_data:
#                                 for message in message_data["messages"]:
#                                     message_id = message.get("id", "UNKNOWN_ID")  # Default if ID is missing
#                                     sender = message.get("from", "UNKNOWN_SENDER")

#                                     if message.get("type") == "text":
#                                         text_body = message["text"].get("body", "")

#                                         # Save to database
#                                         CustomerReply.objects.create(
#                                             sender=sender,
#                                             message_id=message_id,
#                                             text_body=text_body
#                                         )
#                                         logger.info(f"Stored customer reply from {sender}: {text_body}")

#             return JsonResponse({"success": True}, status=200)

#         except json.JSONDecodeError:
#             logger.error("Invalid JSON in the request body")
#             return JsonResponse({"error": "Invalid JSON"}, status=400)


@login_required
@role_required('Accounts')  # Restrict to Accounts role
def account_delivered_orders(request):
    """ View for accounts department to see all delivered DOs and update status """
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', 'all')

    # Base queryset - exclude Received and Cancelled by default
    delivered_orders = DeliveryOrder.objects.exclude(status__in=['Received by A/c', 'Cancelled']).order_by('-date')

    # Apply status filter if not 'all'
    if status_filter != 'all':
        delivered_orders = delivered_orders.filter(status=status_filter)

    # Apply search filter if query exists
    if search_query:
        delivered_orders = delivered_orders.filter(
            Q(do_number__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(salesman__icontains=search_query) |
            Q(driver__icontains=search_query)
        )

    # Get count before pagination
    delivered_count = delivered_orders.count()

    # Pagination
    paginator = Paginator(delivered_orders, 1000)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'orders/account_delivered_orders.html', {
        'orders': page_obj,
        'delivered_count': delivered_count,
        'search_query': search_query,
        'status_filter': status_filter,
        'status_choices': DeliveryOrder.DO_STATUS_CHOICES,
    })



@login_required
@role_required('Accounts')  # Restrict to Accounts role
def mark_received_by_accounts(request, order_id):
    """ View to update DO status to 'Received by A/c' """
    order = get_object_or_404(DeliveryOrder, id=order_id)

    if request.method == 'POST':
        order.status = 'Received by A/c'
        order.received_date = timezone.now()
        order.save()
        return redirect('account_delivered_orders')  # Redirect back to the list

    return render(request, 'orders/confirm_status_change.html', {'order': order})


@login_required
# @role_required('Accounts')  # Restrict to Accounts role
def received_list(request):
    """ View for accounts department to see all delivered DOs and update status """
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', 'all')

    # Base queryset
    orders = DeliveryOrder.objects.filter(status='Received by A/c').order_by('-date')

    # Apply status filter
    if status_filter != 'all':
        orders = orders.filter(status=status_filter)

    # Apply search filter if query exists
    if search_query:
        orders = orders.filter(
            Q(do_number__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(salesman__icontains=search_query) |
            Q(driver__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(orders, 300)  # 300 orders per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Pass the search query and filters back to template
    return render(request, 'orders/received_list.html', {
        'orders': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
    })

    # Pass the search query back to template to maintain it in pagination links
    return render(request, 'orders/received_list.html', {
        'orders': page_obj,
        'search_query': search_query,
    })
from django.shortcuts import render
from .models import CustomerReply, MessageStatus

def messages_dashboard(request):
    customer_replies = CustomerReply.objects.all().order_by('-id')  # Latest first
    message_statuses = MessageStatus.objects.all().order_by('-id')

    dubai_tz = pytz.timezone('Asia/Dubai')

    for status in message_statuses:
        try:
            unix_time = int(status.timestamp)
            utc_time = datetime.utcfromtimestamp(unix_time).replace(tzinfo=pytz.utc)
            status.parsed_time = utc_time.astimezone(dubai_tz)
        except (ValueError, OSError):
            status.parsed_time = None


    return render(request, "orders/messages_dashboard.html", {
        "customer_replies": customer_replies,
        "message_statuses": message_statuses
    })





##############################################################################          TRANSFERS        #############################################################################


@login_required
@role_required('Admin', 'Junaid Admin')
def transfer_list(request):

    # Start with all orders
    transfers = TransferOrder.objects.all().order_by('-date')

    # Filter by date range
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    if from_date and to_date:
        transfers = transfers.filter(date__range=[from_date, to_date])
    elif from_date:
        transfers = transfers.filter(date__gte=from_date)
    elif to_date:
        transfers = transfers.filter(date__lte=to_date)

    # Filter by specific date
    date_filter = request.GET.get('date')
    if date_filter:
        transfers = transfers.filter(date=date_filter)

    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter:
        transfers = transfers.filter(status=status_filter)

    # Filter by driver
    driver_filter = request.GET.get('driver')
    if driver_filter:
        transfers = transfers.filter(driver=driver_filter)

    # Filter by vehicle
    vehicle_filter = request.GET.get('vehicle')
    if vehicle_filter:
        transfers = transfers.filter(vehicle_id=vehicle_filter)

    # Apply search query
    search_query = request.GET.get('search_query')
    if search_query:
        # Filter by DO Number, Customer Name, or Mobile
        transfers = transfers.filter(
            models.Q(t_number__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(transfers, 10)  # Show 10 orders per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)


    # Pass data to the template
    return render(request, 'orders/transfer_list.html', {
        'transfers': page_obj,
        'vehicles': Vehicle.objects.all(),
        'selected_vehicle': vehicle_filter,
        'search_query': search_query,  # Pass the search query back to the template
    })

def transfer_upload_file(request):
    messages = []  # Store error messages

    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.append("❌ Please upload the Excel file!")
            return render(request, 'orders/transfer_upload_file.html', {'messages': messages})

        try:
            excel_file = request.FILES['excel_file']
            fs = FileSystemStorage()
            filename = fs.save(excel_file.name, excel_file)
            file_path = fs.path(filename)

            try:
                with pd.ExcelFile(file_path, engine='openpyxl') as xls:
                    try:
                        df = pd.read_excel(xls, sheet_name='TRANSFERS')
                        df['DO'] = df['DO'].astype(str).str.rstrip('.0')

                        for index, row in df.iterrows():
                            try:
                                TransferOrder.objects.create(
                                    t_number=row['DO'],
                                    date=row['DATE'],
                                    city=row['WAREHOUSE'],
                                )
                            except IntegrityError:
                                messages.append(f"❌ Duplicate DO Number: {row['DO']} already exists! Others are uploaded")
                    except Exception as e:
                        messages.append(f"❌ Error processing 'TRANSFERS' sheet: {str(e)}")

            except Exception as e:
                messages.append(f"❌ Failed to read the Excel file: {str(e)}")

        except Exception as e:
            messages.append(f"❌ Unexpected error while handling file: {str(e)}")

        # Optional: Delete the file after processing
        # if os.path.exists(file_path):
        #     os.remove(file_path)

        if not messages:
            return redirect('transfer_list')

    return render(request, 'orders/transfer_upload_file.html', {'messages': messages})



@login_required
@role_required('Warehouse')
def transfer_update_vehicle(request, vehicle_id):
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)

    if request.method == 'POST':
        # Get the selected DO numbers as a comma-separated string and split it into a list
        t_numbers = request.POST.get('t_numbers', '').split(',')

        # Get the selected driver
        driver = request.POST.get('driver')

        new_status = 'Delivered'
        uae_tz = pytz.timezone('Asia/Dubai')
        # new_date = timezone.now().astimezone(uae_tz) if vehicle and vehicle.id == 17 else None

        TransferOrder.objects.filter(
            vehicle=vehicle
        ).exclude(  # Exclude delivered orders
            status__in=['Delivered']
        ).update(
            vehicle=None,
            status='Pending',
            driver=None
        )

        # Update the selected DOs
        TransferOrder.objects.filter(t_number__in=t_numbers).update(
            vehicle=vehicle,
            status= new_status,
            driver=driver,
        )

        return redirect('vehicle_list')

    # Get all DOs that are not yet delivered and not assigned to any vehicle
    available_orders = TransferOrder.objects.filter(status='Pending')

    # Get DOs already assigned to this vehicle
    assigned_orders = TransferOrder.objects.filter(vehicle=vehicle).exclude(status__in=['Delivered'])

    return render(request, 'orders/transfer_update_vehicle.html', {
        'vehicle': vehicle,
        'available_orders': available_orders,
        'assigned_orders': assigned_orders,
        'drivers': TransferOrder.DRIVER_CHOICES
    })



@login_required
@role_required('Warehouse')
def transfer_vehicle_list(request):
    vehicles = Vehicle.objects.filter(vehicle_number__in=["77507", "99643", "85121", "32438", "72852", "25126","RENTAL-TRANSFER"])
    return render(request, 'orders/transfer_vehicle_list.html', {'vehicles': vehicles})



#####################################################################################  CREDIT PAYMENT SECTION    ###########################################################################'
##################################################################################          ###### #######       ##########################################################################


from .models import Customer,CreditPayment
from django.db.models import Count
from dateutil.relativedelta import relativedelta
from django.db.models.functions import TruncMonth
@login_required
@role_required('Collection', 'Admin', 'Salesman')
def customer_list(request):
    # Fetch the logged‑in user’s role string
    try:
        user_role = request.user.role.role
    except Role.DoesNotExist:
        user_role = None

    query = request.GET.get('q', '').strip()
    salesman_filter = request.GET.get('salesman', '').strip()
    area_filter = request.GET.get('area', '').strip()

    cut_off_date = timezone.now().date() - timedelta(days=100)

    if user_role == 'Salesman':
        # Salesman sees only their own customers, matching deliveryorder.salesman
        name = request.user.first_name
        customers = Customer.objects.filter(
            delivery_orders__salesman__iexact=name
        ).distinct()
        # Still allow searching by name
        if query:
            customers = customers.filter(name__icontains=query)

    else:
        # Collection/Admin: see all, plus dropdown filters
        customers = Customer.objects.all()
        if query:
            customers = customers.filter(name__icontains=query)
        if salesman_filter:
            customers = customers.filter(
                delivery_orders__salesman__iexact=salesman_filter
            ).distinct()
            if salesman_filter.strip().upper() == "MUZAIN":
                customers = customers.exclude(
                    name__in=[
                        "HELIUM ELECTROMICANICAL WORKS L.L.C",
                        "LAMIRADA CONTRACTING LLC",
                        "IAS LOOTAH CONTRACTING"
                    ]
                )

        ########## EVERY MONTH CUSTOMER COUNT ##########

    # today = date.today()

    # # Generate first day of last 3 months
    # month_1 = today.replace(day=1)
    # month_2 = (today - relativedelta(months=1)).replace(day=1)
    # month_3 = (today - relativedelta(months=2)).replace(day=1)

    # # Get end-of-month for each
    # def end_of_month(d):
    #     return (d + relativedelta(months=1)) - timedelta(days=1)

    # # Get active customers per month
    # custs_month_1 = set(
    #     Customer.objects.filter(delivery_orders__date__range=(month_1, end_of_month(month_1))).values_list('id', flat=True)
    # )
    # custs_month_2 = set(
    #     Customer.objects.filter(delivery_orders__date__range=(month_2, end_of_month(month_2))).values_list('id', flat=True)
    # )
    # custs_month_3 = set(
    #     Customer.objects.filter(delivery_orders__date__range=(month_3, end_of_month(month_3))).values_list('id', flat=True)
    # )

    # # Intersect customers who ordered in all 3 months
    # active_every_month_ids = custs_month_1 & custs_month_2 & custs_month_3

    # # Get queryset
    # active_every_month_customers = Customer.objects.filter(id__in=active_every_month_ids)
    # active_every_month_customers_count = active_every_month_customers.count()
    ##### END OF EVERY MONTH CUSTOMER COUNT #####


    if user_role == 'Salesman':
        name = request.user.first_name
        active_customers_count = Customer.objects.filter(
            delivery_orders__salesman=name,
            delivery_orders__date__gte=cut_off_date
        ).distinct().count()
        one_time_customers = Customer.objects.filter(
            delivery_orders__salesman=name,
            delivery_orders__date__gte=cut_off_date
        ).annotate(
            recent_orders=Count('delivery_orders', filter=Q(delivery_orders__date__gte=cut_off_date), distinct=True)
        ).filter(recent_orders=1)
        one_time_customers_count = one_time_customers.count()
    else:
        active_customers_count = Customer.objects.filter(
                delivery_orders__date__gte=cut_off_date
            ).distinct().count()
        one_time_customers = Customer.objects.filter(
            delivery_orders__date__gte=cut_off_date
        ).annotate(
            recent_orders=Count('delivery_orders', filter=Q(delivery_orders__date__gte=cut_off_date), distinct=True)
        ).filter(recent_orders=1)
        one_time_customers_count = one_time_customers.count()
    for cust in one_time_customers:
        orders = cust.delivery_orders.filter(date__gte=cut_off_date)
        print(cust.name, orders.count(), list(orders.values_list('date', flat=True)))
    # print('One Time Customers:', one_time_customers)
    # Prepare stats

    # print("All Month : ",active_every_month_customers)
    total_count = CreditPayment.objects.count()


    if user_role == 'Salesman':
        pending_count = CreditPayment.objects.filter(
        payment_received=False,
        delivery_order__salesman__iexact=request.user.first_name
        ).count()
        total_count = CreditPayment.objects.filter(delivery_order__salesman__iexact=request.user.first_name).count()

    elif user_role == 'Collection':
    # Start with all pending payments
        pending_qs = CreditPayment.objects.filter(payment_received=False)
        total_count_qs = CreditPayment.objects.all()

    # Apply filters if selected
        if salesman_filter:
            pending_qs = pending_qs.filter(delivery_order__salesman__iexact=salesman_filter)
            total_count_qs = total_count_qs.filter(delivery_order__salesman__iexact=salesman_filter)
        if area_filter:
            pending_qs = pending_qs.filter(delivery_order__area__iexact=area_filter)
            total_count_qs = total_count_qs.filter(delivery_order__area__iexact=area_filter)

        pending_count = pending_qs.count()
        total_count = total_count_qs.count()

    else:
        pending_count = CreditPayment.objects.filter(payment_received=False).count()

    # Attach pending count per customer
    for cust in customers:
        cust.pending_counts_customer = (
            CreditPayment.objects
            .filter(delivery_order__customer=cust, payment_received=False)
            .count()
        )

    return render(request, 'credit/customer_list.html', {
        'customers': customers,
        'query': query,
        'active_customers_count': active_customers_count,
        'one_time_customers_count': one_time_customers_count,
        'salesman': salesman_filter,
        'area': area_filter,
        'salesmen': (
            DeliveryOrder.objects
            .exclude(salesman__isnull=True)
            .exclude(salesman='')
            .values_list('salesman', flat=True)
            .distinct()
        ),
        'areas': (
            DeliveryOrder.objects
            .exclude(area__isnull=True)
            .exclude(area='')
            .values_list('area', flat=True)
            .distinct()
        ),
        'pending_count': pending_count,
        'total_count': total_count,
        'user_role': user_role,
    })



@login_required
@role_required('Collection','Salesman')
def customer_credit_entries(request, customer_id, entry_id):
    # Get the specific customer and credit entry
    customer = get_object_or_404(Customer, id=customer_id)
    credit_entry = get_object_or_404(CreditPayment, id=entry_id, delivery_order__customer_id=customer_id)

    # Get all entries for the customer (for context if needed)
    entries = CreditPayment.objects.filter(
        delivery_order__customer=customer,
        delivery_order__status__in=['Delivered', 'Received by A/c']
    ).order_by('-due_date')

    # Calculate exceeded_days for all entries
    for entry in entries:
        entry.exceeded_days = max((date.today() - entry.due_date).days, 0)
        entry.save()

    return render(request, 'credit/customer_entries.html', {
        'customer': customer,
        'entries': entries,
        'credit_entry': credit_entry,
    })


from datetime import date
@login_required
@role_required('Collection','Salesman')
def submit_request_to_md(request, customer_id, entry_id):
    # Verify the entry belongs to the customer
    entry = get_object_or_404(
        CreditPayment,
        id=entry_id,
        delivery_order__customer_id=customer_id
    )

    if request.method == 'POST':
        remark = request.POST.get('remark')
        cheque_date_str = request.POST.get('cheque_date')
        if remark:
            entry.remark = remark

            # entry.customer_cheque_date = date.today()  # Set to today's date
            entry.status_of_approval = 'Pending'  # Set status to Pending for MD review
            entry.save()
            messages.success(request, "Request successfully submitted to MD")

            # Send WhatsApp notification
            whatsapp_message = (
                f"New credit request for approval:\n"
                f"Customer: {entry.delivery_order.customer.name}\n"
                f"Invoice: {entry.delivery_order.invoice_number}\n"
                f"Amount: {entry.delivery_order.amount}\n"
                f"Due Date: {entry.due_date}\n"
                f"Remark: {remark}"
            )
            # Implement your WhatsApp sending logic here

        return redirect('customer_credit_entries',
                      customer_id=customer_id,
                      entry_id=entry_id)

    return redirect('customer_credit_entries',
                  customer_id=customer_id,
                  entry_id=entry_id)

# @login_required
# @role_required('Manager', 'Admin')
# def md_pending_requests(request):
#     entries = CreditPayment.objects.filter(status_of_approval='Pending', remark__isnull=False).order_by('-due_date')

#     overdue_count = entries.filter(exceeded_days__gt=30).count()
#     total_amount = sum(
#         entry.delivery_order.amount for entry in entries
#         if entry.delivery_order and entry.delivery_order.amount
#     )
#     return render(request, 'credit/md_pending_requests.html', {
#         'entries': entries,
#         'overdue_count': overdue_count,
#         'total_amount': total_amount,

#     })

@login_required
@role_required('Manager', 'Admin')
def md_pending_requests(request):
    # 🔹 Get bulk requests
    bulk_requests = CreditBulkRequest.objects.filter(status='Pending') \
                                             .select_related('customer', 'created_by') \
                                             .prefetch_related('creditpayment_set')

    # 🔸 Get individual requests (not in any bulk request)
    individual_requests = CreditPayment.objects.filter(
        status_of_approval='Pending',
         remark__isnull=False,
        bulk_request__isnull=True,
        payment_received=False
    ).select_related('delivery_order__customer')



    return render(request, 'credit/md_pending_requests.html', {
        'bulk_requests': bulk_requests,
        'individual_requests': individual_requests,
    })


@login_required
@role_required('Manager', 'Admin')
def approve_credit_request(request, entry_id):
    entry = get_object_or_404(CreditPayment, id=entry_id)

    if request.method == 'POST':
        decision = request.POST.get('decision')
        if decision == 'approve':
            entry.status_of_approval = 'Approved'
        elif decision == 'decline':
            entry.status_of_approval = 'Declined'
        entry.save()
        return redirect('md_pending_requests')

@login_required
@role_required('Manager', 'Admin')
def approve_bulk_credit_request(request, bulk_id):
    bulk = get_object_or_404(CreditBulkRequest, id=bulk_id)

    if request.method == 'POST':
        decision = request.POST.get('decision')
        if decision not in ['approve', 'decline']:
            messages.error(request, "Invalid decision")
            return redirect('md_pending_requests')

        # Update the bulk status
        bulk.status = 'Approved' if decision == 'approve' else 'Declined'
        bulk.save()

        # Update all associated entries
        CreditPayment.objects.filter(bulk_request=bulk).update(status_of_approval=bulk.status)

        messages.success(request, f"{bulk.status} all entries in Bulk Request #{bulk.id}")
        return redirect('md_pending_requests')



from django.shortcuts import render, redirect, get_object_or_404
from .models import CreditPayment
from django.contrib import messages
from datetime import date
@login_required
@role_required('Collection','Salesman')
def check_cheque_date(request, customer_id, entry_id):
    entry = get_object_or_404(
        CreditPayment,
        id=entry_id,
        delivery_order__customer_id=customer_id
    )

    if request.method == 'POST':
        cheque_date_str = request.POST.get('cheque_date')

        if cheque_date_str:
            try:
                cheque_date = date.fromisoformat(cheque_date_str)
                entry.customer_cheque_date = cheque_date

                # Important: Compare and update approval status here!
                if cheque_date <= entry.due_date:
                    entry.status_of_approval = 'Approved'
                else:
                    entry.status_of_approval = 'Pending Approval'

                entry.save()

                # You can also return if approved or pending in JSON
                return JsonResponse({
                    'success': True,
                    'approved': cheque_date <= entry.due_date,
                    'message': 'Cheque date saved successfully'
                })

            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid date format'
                }, status=400)

        return JsonResponse({
            'success': False,
            'message': 'No cheque date provided'
        }, status=400)

    return JsonResponse({
        'success': False,
        'message': 'Invalid request method'
    }, status=405)

from django.contrib import messages
@login_required
@role_required('Collection','Salesman')
def mark_payment_received(request, customer_id, entry_id):
    # Verify the entry belongs to the customer
    entry = get_object_or_404(
        CreditPayment,
        id=entry_id,
        delivery_order__customer_id=customer_id
    )

    if request.method == 'POST':
        entry.payment_received = True
        entry.save()
        messages.success(request, "Payment successfully marked as received")
    return redirect('customer_credit_entries',
                      customer_id=customer_id,
                      entry_id=entry_id)





@login_required
@role_required('Manager', 'Admin')
def md_customer_entries(request):
    customers = Customer.objects.all()
    entries = None

    selected_customer = None

    if request.method == 'POST':
        customer_id = request.POST.get('customer_id')
        selected_customer = get_object_or_404(Customer, id=customer_id)
        entries = CreditPayment.objects.filter(delivery_order__customer=selected_customer).order_by('-due_date')

    return render(request, 'credit/md_customer_entries.html', {
        'customers': customers,
        'entries': entries,
        'selected_customer': selected_customer,
    })


# for updating credit limits
from django import forms

class CreditLimitForm(forms.ModelForm):
    class Meta:
        model = Customer
        fields = ['credit_limit', 'credit_limit_amount']
@login_required
@role_required('Manager', 'Admin')
def customer_credit_list_and_update(request):
    customers = Customer.objects.all()

    if request.method == 'POST':
        customer_id = request.POST.get('customer_id')
        customer = get_object_or_404(Customer, id=customer_id)
        form = CreditLimitForm(request.POST, instance=customer)

        if form.is_valid():
            form.save()

            # Update due dates
            for payment in CreditPayment.objects.filter(delivery_order__customer=customer):
                payment.due_date = payment.delivery_order.date + timedelta(days=customer.credit_limit or 120)
                payment.save()

            return redirect('customer_credit_list')  # Or the same view name

    customer_forms = [(customer, CreditLimitForm(instance=customer)) for customer in customers]
    return render(request, 'credit/customer_credit_list.html', {
    'customer_forms': customer_forms,
})

@login_required
@role_required('Collection','Salesman')
def payment_status_by_customer(request, customer_id):
    # Get filter parameters
    selected_month = request.GET.get('month')
    selected_year = request.GET.get('year')

    # Default to current year if not selected
    today = date.today()
    selected_year = int(selected_year) if selected_year else today.year
    selected_month = int(selected_month) if selected_month else None

    # Base queryset with year filter
    base_query = CreditPayment.objects.filter(
        delivery_order__customer_id=customer_id,
        delivery_order__status__in=['Delivered', 'Received by A/c'],
        delivery_order__date__year=selected_year
    )

    # Apply month filter if selected
    if selected_month:
        base_query = base_query.filter(delivery_order__date__month=selected_month)

    # Split into pending/completed
    pending = base_query.filter(payment_received=False)
    completed = base_query.filter(payment_received=True)

    # Calculate totals
    pending_total = pending.aggregate(Sum('delivery_order__amount'))['delivery_order__amount__sum'] or 0
    completed_total = completed.aggregate(Sum('delivery_order__amount'))['delivery_order__amount__sum'] or 0
    credit_note_total =pending.aggregate(Sum('delivery_order__credit_note'))['delivery_order__credit_note__sum'] or 0
    customer = Customer.objects.get(id=customer_id)
    years = range(2020, date.today().year + 1)
    months = list(range(1, 13))

    return render(request, 'credit/payment_status.html', {
        'pending_payments': pending,
        'completed_payments': completed,
        'pending_total': pending_total,
        'credit_note_total':credit_note_total,
        'completed_total': completed_total,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'years': years,
        'pending_count': pending.count(),
        'completed_count': completed.count(),
        'customer': customer,
        'months': months,
        'show_all_months': not bool(selected_month),
    })

from django.db.models import Sum
@login_required
@role_required('Manager', 'Admin')
def md_dashboard(request):
    unpaid_payments = CreditPayment.objects.filter(payment_received=False)
    unpaid_payments_count = unpaid_payments.count()
    paid_payments = CreditPayment.objects.filter(payment_received=True).count()
    overdue_payments_count = unpaid_payments.filter(exceeded_days__gt=30).count()
    total_customer_count = Customer.objects.count()
    total_entries_count = CreditPayment.objects.count()
    do_filter = DeliveryOrder.objects.exclude(status='Cancelled')
    payments = CreditPayment.objects.select_related('delivery_order', 'delivery_order__customer').order_by('-delivery_order__date')



    # Calculate total amount to be received
    # total_amount = unpaid_payments.aggregate(
    #     total=Sum('delivery_order__amount')
    # )['total'] or 0
    total_do_amount = do_filter.aggregate(total=Sum('amount'))['total'] or 0
    delivered_received_amount = payments.filter(payment_received=True).aggregate(Sum('delivery_order__amount'))['delivery_order__amount__sum'] or 0
    total_effective_due = total_effective_due = round(total_do_amount - delivered_received_amount, 2)

    total_amount = unpaid_payments.exclude(
    ~Q(delivery_order__status__in=['Delivered', 'Received by A/c'])
        ).aggregate(
        total=Sum('delivery_order__amount')
        )['total'] or 0

    return render(request, 'credit/md_dashboard.html',{
        'unpaid_payments_count': unpaid_payments_count,
        'paid_payments': paid_payments,
        'total_effective_due': total_effective_due,
        'total_amount': total_amount,
        'overdue_payments_count': overdue_payments_count,
        'total_customer_count': total_customer_count,
        'total_entries_count': total_entries_count,
    })



def password_gate(request):
    # If correct password already entered (stored in session)
    if request.session.get('password_verified', False):
        return redirect('md_dashboard')  # Replace with your actual destination URL name

    if request.method == 'POST':
        password = request.POST.get('password')
        # Replace 'your_secret_password' with your actual password
        if password == 'junaid1023':
            request.session['password_verified'] = True
            return redirect('md_dashboard')  # Replace with your actual destination URL name
        else:
            messages.error(request, 'Incorrect password')

    return render(request, 'credit/password_gate.html')

import uuid
from decimal import Decimal, InvalidOperation

def upload_invoices(request):
    if request.method == 'POST' and request.FILES.get('invoice_file'):
        excel_file = request.FILES['invoice_file']
        fs = FileSystemStorage()
        filename = fs.save(excel_file.name, excel_file)
        file_path = fs.path(filename)

        try:
            df = pd.read_excel(file_path, sheet_name='INVOICES')
            df['DO'] = df['DO'].apply(lambda x: str(int(x)) if isinstance(x, float) and x.is_integer() else str(x))
            df['INVOICE'] = df['INVOICE'].apply(lambda x: str(int(x)) if isinstance(x, float) and x.is_integer() else str(x))

            df['INVOICE'] = df['INVOICE'].apply(lambda x: f"C{uuid.uuid4().hex[:12].upper()}" if str(x).strip().upper() == 'NIL' else x)

            for index, row in df.iterrows():
                try:
                    do_number = str(row.get('DO')).strip()
                    invoice_number = str(row.get('INVOICE')).strip()

                    # Clean and convert amount
                    raw_amount = str(row.get('AMOUNT')).replace('“', '').replace('”', '').replace(',', '').strip()
                    try:
                        amount = Decimal(raw_amount)
                    except InvalidOperation:
                        messages.warning(request, f"❌ Row {index+2}: Invalid amount '{raw_amount}'. Skipped.")
                        continue

                    # Update the delivery order
                    try:
                        order = DeliveryOrder.objects.get(do_number=do_number)

                        # keep a copy of the old invoice before we touch it
                        old_invoice = order.invoice_number

                        # if the order already had an invoice *and* it is about to change,
                        # remove any linked credit‑payment records
                        if old_invoice and old_invoice != invoice_number:
                            CreditPayment.objects.filter(delivery_order=order).delete()

                        # now store the new details
                        order.invoice_number = invoice_number or None
                        order.amount = amount
                        order.save()

                    except DeliveryOrder.DoesNotExist:
                        messages.warning(request, f"❌ Row {index+2}: DO '{do_number}' not found. Skipped.")
                        continue

                except Exception as e:
                    messages.warning(request, f"❌ Row {index+2}: Unexpected error — {e}")
                    continue

            messages.success(request, "✅ Invoice numbers and amounts updated successfully.")
        except Exception as e:
            messages.error(request, f"⚠️ Error processing file: {str(e)}")

        return redirect('order_list')

    return render(request, 'orders/upload_invoices.html')


import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Customer

def export_customer_names(request):
    # Create a workbook and select active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Names"

    # Write the header
    ws['A1'] = 'CUSTOMER CODE'
    ws['B1'] = 'CUSTOMER NAME'
    ws['C1'] = 'PAYMENT TERMS'
    ws['D1'] = 'CREDIT LIMIT'
    ws['E1'] = 'ADDITIONAL TERMS'
    ws['F1'] = 'Monthly Billing'
    # Write data
    for idx, customer in enumerate(Customer.objects.all(), start=2):
        ws[f'A{idx}'] = customer.customer_code
        ws[f'B{idx}'] = customer.name

    # Adjust column width
    ws.column_dimensions[get_column_letter(1)].width = 30
    ws.column_dimensions[get_column_letter(1)].width = 30

    # Create HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=customer_names.xlsx'
    wb.save(response)
    return response



import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Customer

def upload_customer_limits(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']

        try:
            # Read the uploaded Excel file
            df = pd.read_excel(file)
            df

            for _, row in df.iterrows():
                customer_code = str(row.get('CUSTOMER CODE')).strip()
                name = str(row.get('CUSTOMER NAME')).strip()
                payment_terms = row.get('PAYMENT TERMS')
                credit_limit_amount = row.get('CREDIT LIMIT')
                additional_terms = row.get('ADDITIONAL TERMS')
                use_next_month = row.get('Monthly Billing')
                opening_balance = row.get('OPENING BALANCE')

                print(f"Processing: {name}, Terms: {payment_terms}, Limit: {credit_limit_amount}, Monthly: {use_next_month}")

                if name and pd.notnull(payment_terms) and pd.notnull(credit_limit_amount):
                    customer, created = Customer.objects.get_or_create(customer_code=customer_code, defaults={'name': name})
                    customer.credit_limit = int(payment_terms)
                    customer.credit_limit_amount = round(float(credit_limit_amount), 2)
                    customer.additional_terms = additional_terms if pd.notnull(additional_terms) else None
                    customer.opening_balance=opening_balance

                    if pd.notnull(use_next_month):
                        customer.use_next_month_start = str(use_next_month).strip().lower() in ['true', 'yes', '1']
                    else:
                        customer.use_next_month_start = False

                    customer.save()
                    print(f"✅ {'Created' if created else 'Updated'}: {customer.name}")
                else:
                    print(f"❌ Skipped row: Missing required data")
        except Exception as e:
            messages.error(request, f"Error: {e}")

        return redirect('customer_credit_list')  # Redirect to the customer list page

    return render(request, 'credit/upload_customers.html')






from django.db.models import Q, Sum
from django.shortcuts import render
from .models import CreditPayment, Customer
from django.utils.timezone import now

def credit_dashboard(request):
    query = request.GET.get('q', '')
    payment_status = request.GET.get('payment_received')
    approval_status = request.GET.get('approval_status')
    customer_id = request.GET.get('customer')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Get page number from request
    page_number = request.GET.get('page', 1)

    payments = CreditPayment.objects.select_related('delivery_order', 'delivery_order__customer').order_by('-delivery_order__date')

    # Apply filters (your existing filter code remains the same)
    if query:
        payments = payments.filter(
            Q(delivery_order__customer_name__icontains=query) |
            Q(delivery_order__do_number__icontains=query) |
            Q(delivery_order__invoice_number__icontains=query)
        )

    if customer_id:
        payments = payments.filter(delivery_order__customer_id=customer_id)

    if payment_status == 'yes':
        payments = payments.filter(payment_received=True)
    elif payment_status == 'no':
        payments = payments.filter(payment_received=False)

    if approval_status:
        payments = payments.filter(status_of_approval=approval_status)

    if start_date and end_date:
        payments = payments.filter(delivery_order__date__range=[start_date, end_date])
    elif start_date:
        payments = payments.filter(delivery_order__date__gte=start_date)
    elif end_date:
        payments = payments.filter(delivery_order__date__lte=end_date)

    # Create paginator
    paginator = Paginator(payments, 500)  # Show 25 records per page
    page_obj = paginator.get_page(page_number)

    # Your existing amount calculations remain the same
    do_filter = DeliveryOrder.objects.exclude(status='Cancelled')

    if query:
        do_filter = do_filter.filter(
            Q(customer_name__icontains=query) |
            Q(do_number__icontains=query) |
            Q(invoice_number__icontains=query)
        )

    if customer_id:
        do_filter = do_filter.filter(customer_id=customer_id)

    if start_date and end_date:
        do_filter = do_filter.filter(date__range=[start_date, end_date])
    elif start_date:
        do_filter = do_filter.filter(date__gte=start_date)
    elif end_date:
        do_filter = do_filter.filter(date__lte=end_date)

    total_do_amount = do_filter.aggregate(total=Sum('amount'))['total'] or 0


    total_invoiced = payments.filter(
        delivery_order__status__in=['Delivered', 'Received by A/c']
    ).aggregate(total=Sum('delivery_order__amount'))['total'] or 0



    delivered_received_amount = payments.filter(payment_received=True).aggregate(Sum('delivery_order__amount'))['delivery_order__amount__sum'] or 0

    delivered_pending_amount = payments.filter(payment_received=False).aggregate(
    Sum('delivery_order__amount'))['delivery_order__amount__sum'] or 0

    not_delivered_amount = do_filter.exclude(status__in=['Delivered', 'Received by A/c']).aggregate(
        total=Sum('amount')
    )['total'] or 0
    total_effective_due = delivered_pending_amount + not_delivered_amount

    # total_effective_due = total_do_amount - delivered_received_amount
    total_received = payments.filter(payment_received=True).aggregate(Sum('delivery_order__amount'))['delivery_order__amount__sum'] or 0
    total_credit_note=payments.filter(payment_received=False).aggregate(Sum('delivery_order__credit_note'))['delivery_order__credit_note__sum'] or 0

    customers = Customer.objects.all()
    if request.GET.get('export') == 'excel':
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Credit Payments'

        # Header row
        headers = [
            'Customer Name', 'DO Number', 'Invoice Number', 'Date',
            'Amount', 'Due Date', 'Exceeded Days', 'Approval Status',
            'Payment Received', 'Customer Cheque Date'
        ]
        sheet.append(headers)

        for payment in payments:
            do = payment.delivery_order
            row = [
                do.customer_name,
                do.do_number,
                do.invoice_number,
                do.date.strftime('%Y-%m-%d') if do.date else '',
                do.amount,
                payment.due_date.strftime('%Y-%m-%d') if payment.due_date else '',
                payment.exceeded_days,
                payment.status_of_approval,
                'Yes' if payment.payment_received else 'No',
                payment.customer_cheque_date.strftime('%Y-%m-%d') if payment.customer_cheque_date else ''
            ]
            sheet.append(row)

        # Prepare response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=credit_payments.xlsx'
        workbook.save(response)
        return response

    context = {
        'payments': page_obj,  # Changed from payments to page_obj
        'total_do_amount': total_do_amount,
        'total_credit_note':total_credit_note,
        'total_invoiced': total_invoiced,
        'total_received': total_received,
        'total_effective_due': total_effective_due,
        'customers': customers,
        'page_obj': page_obj,  # Add page_obj to context
    }

    return render(request, 'credit/credit_dashboard.html', context)

######################################################          COMBINED CREDIT PAYMENT       ####################################################################################

from django.views.decorators.http import require_POST
from django.db.models import F
from django.views.decorators.http import require_POST

@login_required
def combined_customer_entries(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)

    if request.method == 'POST':
        selected_entry_ids = [int(id) for id in request.POST.getlist('entry_ids')]
        request.session['selected_entry_ids'] = selected_entry_ids

        for entry_id in selected_entry_ids:
            due_date_str = request.POST.get(f'due_date_{entry_id}')
            if due_date_str:
                try:
                    new_due_date = date.fromisoformat(due_date_str)
                    CreditPayment.objects.filter(id=entry_id).update(due_date=new_due_date)
                except ValueError:
                    pass
    else:
        selected_entry_ids = request.session.get('selected_entry_ids', [])

    # 🔴 Move this below entries definition!
    # print(f"Processing {len(entries)} entries")

    # Fetch entries for the customer
    entries = CreditPayment.objects.filter(
        id__in=selected_entry_ids,
        delivery_order__customer=customer
    )

    # Auto-approve logic
    entries.filter(
        customer_cheque_date__isnull=False,
        customer_cheque_date__lte=F('due_date')
    ).update(status_of_approval='Approved')

    # Refresh entries
    entries = CreditPayment.objects.filter(
        id__in=selected_entry_ids,
        delivery_order__customer=customer
    )

    # # 💡 NOW it’s safe to print!
    # print(f"Processing {len(entries)} entries")
    # for i, entry in enumerate(entries):
    #     cheque_info = f"Cheque: {entry.customer_cheque_date}" if entry.customer_cheque_date else "No cheque date"
    #     status_info = f"Status: {entry.status_of_approval}"
    #     print(f"Entry {i+1}: {cheque_info}, Due: {entry.due_date}, {status_info}")

    # Calculate total
    total_amount = round(entries.aggregate(
        total=Sum('delivery_order__amount')
    )['total'] or 0,2)

    # Update exceeded_days
    for entry in entries:
        entry.exceeded_days = max((date.today() - entry.due_date).days, 0)

    # Show submit to MD and mark paid flags
    show_submit_to_md = any(
        entry.customer_cheque_date and entry.customer_cheque_date > entry.due_date
        for entry in entries
    )
    show_mark_paid = all(
        entry.customer_cheque_date and entry.customer_cheque_date <= entry.due_date
        for entry in entries
    )

    return render(request, 'credit/combined_entries.html', {
        'customer': customer,
        'entries': entries,
        'total_amount': total_amount,
        'selected_entry_ids': selected_entry_ids,
        'show_submit_to_md': show_submit_to_md,
        'show_mark_paid': show_mark_paid,
    })


from django.http import JsonResponse
from datetime import date

@login_required
@role_required('Collection', 'Salesman')
@require_POST
def bulk_update_cheque_dates(request, customer_id):

    customer = get_object_or_404(Customer, id=customer_id)


    # 1) Get the entry IDs as a comma-separated string
    entry_ids_str = request.POST.get('entry_ids', '')


    if not entry_ids_str:
        messages.error(request, "No entries selected")
        return redirect('combined_customer_entries', customer_id=customer_id)

    # Split the string into individual IDs
    try:
        entry_ids = [int(eid) for eid in entry_ids_str.split(',') if eid]
    except ValueError:
        messages.error(request, "Invalid entry ID format")
        return redirect('combined_customer_entries', customer_id=customer_id)

    print(f"Parsed entry IDs: {entry_ids}")

    # 2) Get the cheque date
    cheque_date_str = request.POST.get('cheque_date')
    print(f"Received cheque date: {cheque_date_str}")

    if not cheque_date_str:
        messages.error(request, "No cheque date provided")
        return redirect('combined_customer_entries', customer_id=customer_id)

    try:
        customer_cheque_date = date.fromisoformat(cheque_date_str)
        print(f"Parsed cheque date: {customer_cheque_date}")
    except ValueError:
        messages.error(request, "Invalid date format")
        return redirect('combined_customer_entries', customer_id=customer_id)

    # 3) Fetch and update entries
    entries = CreditPayment.objects.filter(
        id__in=entry_ids,
        delivery_order__customer=customer
    )

    print(f"Found {entries.count()} entries to update")

    if not entries.exists():
        messages.error(request, "No valid entries found")
        return redirect('combined_customer_entries', customer_id=customer_id)

    updated_count = 0
    for entry in entries:
        entry.customer_cheque_date = customer_cheque_date
        entry.status_of_approval = 'Approved' if customer_cheque_date <= entry.due_date else 'Pending'
        entry.save()
        updated_count += 1



    messages.success(
        request,
        f"Succesfully Updated cheque date for {updated_count} entr{'y' if updated_count == 1 else 'ies'}"
    )
    print(f"Successfully updated {updated_count} entries")
    return redirect('combined_customer_entries', customer_id=customer_id)



@login_required
@role_required('Collection','Salesman')
@require_POST
def bulk_mark_paid(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    entry_ids = request.POST.get('entry_ids', '').split(',')

    # Get all valid entries that belong to this customer
    entries = CreditPayment.objects.filter(
        id__in=entry_ids,
        delivery_order__customer=customer
    )

    marked_count = 0
    for entry in entries:
        if not entry.payment_received:
            entry.payment_received = True
            entry.save()
            marked_count += 1

    messages.success(request, f"Marked {marked_count} entries as paid")
    return redirect('combined_customer_entries', customer_id=customer_id)



@login_required
@role_required('Collection','Salesman')
@require_POST
def bulk_submit_to_md(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    entry_ids = request.POST.get('entry_ids', '').split(',')
    remark = request.POST.get('remark', '').strip()

    if not remark:
        messages.error(request, "Please provide remarks for MD approval")
        return redirect('combined_customer_entries', customer_id=customer_id)

    entries = CreditPayment.objects.filter(
        id__in=entry_ids,
        delivery_order__customer=customer
    ).exclude(payment_received=True)

    if not entries.exists():
        messages.error(request, "No valid entries to submit.")
        return redirect('combined_customer_entries', customer_id=customer_id)

    # Create the bulk request
    bulk_request = CreditBulkRequest.objects.create(
        customer=customer,
        created_by=request.user,
        remark=remark
    )
    print("Bulk submit triggered for customer:", customer.name)
    submitted_count = 0
    invoice_numbers = []

    for entry in entries:
        entry.remark = remark
        entry.status_of_approval = 'Pending'
        entry.bulk_request = bulk_request
        entry.save()
        submitted_count += 1
        invoice_numbers.append(entry.delivery_order.invoice_number)

    # Optional: send WhatsApp message here

    messages.success(request, f"Submitted {submitted_count} entries to MD as one bulk request.")
    return redirect('combined_customer_entries', customer_id=customer_id)


# @login_required
# @role_required('Collection','Salesman')
# def select_entries_for_combined_view(request, customer_id):
#     customer = get_object_or_404(Customer, id=customer_id)

#     # Get pending payments for this customer
#     pending_payments = CreditPayment.objects.filter(
#         delivery_order__customer=customer,
#         payment_received=False
#     ).select_related('delivery_order')

#     return render(request, 'credit/select_entries.html', {
#         'customer': customer,
#         'pending_payments': pending_payments,
#     })

@login_required
@role_required('Collection','Salesman')
def select_entries_for_combined_view(request, customer_id):
    # Get filter parameters
    selected_month = request.GET.get('month')
    selected_year = request.GET.get('year')

    # Default to current year if not selected
    today = date.today()
    selected_year = int(selected_year) if selected_year else today.year
    selected_month = int(selected_month) if selected_month else None

    # Base queryset
    base_query = CreditPayment.objects.filter(
        delivery_order__customer_id=customer_id,
        payment_received=False
    ).exclude(
        delivery_order__status='Cancelled'
    )

    # Apply year filter
    base_query = base_query.filter(delivery_order__date__year=selected_year)

    # Apply month filter if selected
    if selected_month:
        base_query = base_query.filter(delivery_order__date__month=selected_month)

    # Calculate total
    pending_total = base_query.aggregate(Sum('delivery_order__amount'))['delivery_order__amount__sum'] or 0

    customer = get_object_or_404(Customer, id=customer_id)
    years = range(2020, date.today().year + 1)
    months = list(range(1, 13))

    return render(request, 'credit/select_entries.html', {
        'customer': customer,
        'pending_payments': base_query.select_related('delivery_order'),
        'selected_month': selected_month,
        'selected_year': selected_year,
        'years': years,
        'months': months,
        'pending_total': pending_total
    })


from django.shortcuts import render
from django.http import HttpResponse
from django.contrib import messages
from openpyxl import load_workbook
from .models import DeliveryOrder
def bulk_upload_credit_notes(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        try:
            wb = load_workbook(filename=excel_file)
            sheet = wb.active

            success_count = 0
            error_count = 0
            error_messages = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) < 2:
                    continue

                invoice_number, credit_note_amount = row[0], row[1]

                try:
                    # Skip if empty values
                    if not invoice_number or credit_note_amount is None:
                        continue

                    # Convert to Decimal (consistent with model field type)
                    from decimal import Decimal
                    credit_note_amount = Decimal(str(credit_note_amount))

                    do = DeliveryOrder.objects.get(invoice_number=str(invoice_number))

                    # Ensure amount exists before calculation
                    if do.amount is None:
                        error_messages.append(f"Invoice {invoice_number} has no amount specified")
                        error_count += 1
                        continue

                    do.credit_note = credit_note_amount
                    do.amount_after_credit_note = do.amount - do.credit_note
                    do.save()
                    success_count += 1

                except DeliveryOrder.DoesNotExist:
                    error_messages.append(f"Invoice {invoice_number} not found")
                    error_count += 1
                except ValueError as e:
                    error_messages.append(f"Invalid amount format for invoice {invoice_number}: {str(e)}")
                    error_count += 1
                except Exception as e:
                    error_messages.append(f"Error processing {invoice_number}: {str(e)}")
                    error_count += 1

            messages.success(request, f"Successfully updated {success_count} credit notes")
            if error_count > 0:
                messages.warning(request, f"{error_count} errors occurred")

            # Store errors in session if too many
            if len(error_messages) > 10:
                request.session['bulk_upload_errors'] = error_messages
                messages.info(request, "See detailed errors in download below")

            return redirect('bulk_upload_credit_notes')

        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
            return redirect('bulk_upload_credit_notes')

    # For GET requests
    errors = request.session.pop('bulk_upload_errors', [])
    return render(request, 'credit/bulk_upload.html', {'errors': errors})




################################  MANUAL ENTER ##########################
from .forms import PreEnteredDOBulkForm
from .models import PreEnteredDO
from django.contrib import messages
import re

def enter_do_number(request):
    if request.method == 'POST':
        form = PreEnteredDOBulkForm(request.POST)
        if form.is_valid():
            raw_input = form.cleaned_data['do_numbers']
            # Split on commas, newlines, and spaces
            do_list = [item.strip() for item in re.split(r'[,\n\s]+', raw_input) if item.strip()]

            saved = 0
            skipped = 0
            for do_num in do_list:
                if not PreEnteredDO.objects.filter(do_number=do_num).exists():
                    PreEnteredDO.objects.create(do_number=do_num)
                    saved += 1
                else:
                    skipped += 1

            messages.success(request, f"✅ Saved: {saved} DO number(s). Skipped: {skipped} already exist.")
            return redirect('enter_do_number')
    else:
        form = PreEnteredDOBulkForm()
    recent_entries = PreEnteredDO.objects.order_by('-entered_at')[:5]
    print(recent_entries)
    return render(request, 'orders/enter_do_number.html', {'form': form, 'recent_entries': recent_entries})


def entered_do_history(request):
    # Get the latest 5 entries, ordered by most recent
    recent_entries = PreEnteredDO.objects.order_by('-id')
    return render(request, 'orders/entered_do_history.html', {'recent_entries': recent_entries})

from django.http import JsonResponse


def all_orders_json(request):
    data = list(DeliveryOrder.objects.all().values(
        'do_number', 'status', 'date', 'customer_name','mobile_number'
    ))
    return JsonResponse(data, safe=False)






def refresh_customer_stats(request):
    today = date.today()

    # Generate first day of last 3 months
    month_1 = today.replace(day=1)
    month_2 = (today - relativedelta(months=1)).replace(day=1)
    month_3 = (today - relativedelta(months=2)).replace(day=1)

    # Get end-of-month for each
    def end_of_month(d):
        return (d + relativedelta(months=1)) - timedelta(days=1)

    # Get active customers per month
    custs_month_1 = set(
        Customer.objects.filter(delivery_orders__date__range=(month_1, end_of_month(month_1))).values_list('id', flat=True)
    )
    custs_month_2 = set(
        Customer.objects.filter(delivery_orders__date__range=(month_2, end_of_month(month_2))).values_list('id', flat=True)
    )
    custs_month_3 = set(
        Customer.objects.filter(delivery_orders__date__range=(month_3, end_of_month(month_3))).values_list('id', flat=True)
    )

    # Intersect customers who ordered in all 3 months
    active_every_month_ids = custs_month_1 & custs_month_2 & custs_month_3

    # Get queryset
    active_every_month_customers = Customer.objects.filter(id__in=active_every_month_ids)
    active_every_month_customers_count = active_every_month_customers.count()

    # Calculate intersect
    customer_ids = list(custs_month_1 & custs_month_2 & custs_month_3)
    update_customer_frequencies()
    # Save result to DB
    CachedCustomerStats.objects.update_or_create(
        id=1,
        defaults={
            'count_every_month': len(customer_ids),
            'customer_ids': customer_ids,
        }
    )

    return JsonResponse({'status': 'success', 'count': len(customer_ids)})



from datetime import date, timedelta
from dateutil.relativedelta import relativedelta
from .models import Customer, DeliveryOrder

def update_customer_frequencies():
    today = date.today()
    start_date = (today - relativedelta(months=2)).replace(day=1)  # start of month 3 ago
    end_date = today

    # Get frequency counts for each customer
    freq_dict = (
        DeliveryOrder.objects
        .filter(date__range=(start_date, end_date))
        .values('customer_id')
        .annotate(order_count=models.Count('id'))
    )

    # Reset all to 0 first (optional, if you want accurate counts)
    Customer.objects.update(frequency_last_3_months=0)

    # Update each customer with new frequency
    for entry in freq_dict:
        Customer.objects.filter(id=entry['customer_id']).update(frequency_last_3_months=entry['order_count'])





from django.http import JsonResponse
from .models import MessageStatus

def message_status_list(request):
    data = list(MessageStatus.objects.all().values(
        'message_id', 'recipient_id', 'status', 'timestamp'
    ))
    return JsonResponse(data, safe=False)




from django.db.models.functions import TruncMonth
from django.db.models import Count, Sum
from collections import defaultdict
import calendar
from datetime import datetime
from dateutil.relativedelta import relativedelta

@login_required

def customer_frequency_analysis(request):
    # Get period (default: last 6 months)
    start_month_str = request.GET.get('start')
    end_month_str = request.GET.get('end')
    today = timezone.now().date()

    if start_month_str and end_month_str:
        start_date = datetime.strptime(start_month_str, "%Y-%m").date().replace(day=1)
        end_date = datetime.strptime(end_month_str, "%Y-%m").date().replace(day=1)
        last_day = calendar.monthrange(end_date.year, end_date.month)[1]
        end_date = end_date.replace(day=last_day)
    else:
        end_date = today
        start_date = (today.replace(day=1) - relativedelta(months=5))

    # Salesman filter
    salesman_filter = request.GET.get('salesman', '').strip()

    # Fetch delivery orders (exclude cancelled)
    delivery_orders = DeliveryOrder.objects.filter(
        date__range=[start_date, end_date]
    ).exclude(status__iexact='Cancelled')

    if salesman_filter:
        delivery_orders = delivery_orders.filter(salesman__iexact=salesman_filter)

    # Determine all months in range
    total_months = set(
        delivery_orders
        .annotate(month=TruncMonth('date'))
        .values_list('month', flat=True)
        .distinct()
    )
    total_months_count = len(total_months)

    # Collect orders per customer
    customer_data = defaultdict(lambda: {
        "months": set(),
        "orders": 0,
        "salesman": "",
        "name": "",
        "total_value": 0.0
    })

    for order in delivery_orders.annotate(month=TruncMonth('date')):
        cust_id = order.customer.id
        customer_data[cust_id]["name"] = order.customer.name
        customer_data[cust_id]["salesman"] = order.salesman
        customer_data[cust_id]["orders"] += 1
        customer_data[cust_id]["months"].add(order.month)
        customer_data[cust_id]["total_value"] += float(order.amount or 0)  # 👈 add order value

    # Initialize stats
    stats = {
        "one_month": 0,
        "two_month": 0,
        "all_month": 0,
        "one_time": 0,
        "two_time": 0,
        "total_value": 0.0,
    }

    results = []
    for cust_id, info in customer_data.items():
        order_count = info["orders"]
        month_count = len(info["months"])
        total_value = info["total_value"]

        # Order classification
        if order_count == 1:
            stats["one_time"] += 1
            order_class = "One-Time Customer"
        elif order_count == 2:
            stats["two_time"] += 1
            order_class = "Two-Time Customer"
        else:
            order_class = f"{order_count} Orders"

        # Month classification
        if month_count == 1:
            stats["one_month"] += 1
            month_class = "One-Month Customer"
        elif month_count == 2:
            stats["two_month"] += 1
            month_class = "Two-Month Customer"
        elif month_count == total_months_count:
            stats["all_month"] += 1
            month_class = "All-Month Customer"
        else:
            month_class = f"{month_count} Months"

        stats["total_value"] += total_value  # 👈 accumulate overall total

        results.append({
            "id": cust_id,
            "name": info["name"],
            "salesman": info["salesman"],
            "orders": order_count,
            "months": [m.strftime("%b-%Y") for m in sorted(info["months"], key=lambda d: d)],
            "order_class": order_class,
            "month_class": month_class,
            "total_value": round(total_value, 2)
        })

    return render(request, 'credit/customer_frequency_analysis.html', {
        "results": results,
        "stats": stats,
        "salesmen": (
            DeliveryOrder.objects.exclude(salesman__isnull=True)
            .exclude(salesman="")
            .values_list("salesman", flat=True)
            .distinct()
        ),
        "selected_salesman": salesman_filter,
        "start": start_date.strftime("%Y-%m"),
        "end": end_date.strftime("%Y-%m"),
        "total_months": sorted([m.strftime("%b-%Y") for m in total_months])
    })





# --- NEW imports at top ---
import io
from decimal import Decimal, InvalidOperation
from datetime import datetime
import pandas as pd

from django.contrib import messages
from django.shortcuts import render, redirect
from django.db import transaction
from django.http import HttpResponse
from django.db.models import Count, Sum, Value
from django.db.models.functions import TruncMonth, Coalesce

from .forms import SAPInvoiceUploadForm
from .models import SAPInvoice, SAPInvoiceUploadBatch

# ---------- Helpers ----------
def _coerce_decimal(x):
    if pd.isna(x):
        return Decimal("0.00")
    if isinstance(x, (int, float)):
        return Decimal(str(x))
    s = str(x).replace(",", "").strip()
    if s == "":
        return Decimal("0.00")
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal("0.00")

def _read_sap_dataframe(uploaded_file):
    # Import inside to avoid GET crashes if pandas/openpyxl aren't installed
    import pandas as pd
    from decimal import Decimal, InvalidOperation
    from datetime import datetime

    def _coerce_decimal(x):
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return Decimal("0.00")
        s = str(x).replace(",", "").strip()
        if not s:
            return Decimal("0.00")
        try:
            return Decimal(s)
        except InvalidOperation:
            return Decimal("0.00")

    # 1) Read a small sample to find the header row (where expected column names appear)
    probe = pd.read_excel(uploaded_file, header=None, nrows=20, dtype=str, engine="openpyxl")
    probe = probe.applymap(lambda v: v.strip() if isinstance(v, str) else v)

    expected_any = {"Date", "Customer Name", "Sales Employee", "Cancelled", "Document Total", "#"}
    header_row = None
    for i in range(min(20, len(probe))):
        row_vals = set(str(v).strip() for v in probe.iloc[i].tolist() if v is not None and str(v).strip() != "")
        # if at least 3 expected tokens found, assume this is the header
        if len(expected_any.intersection(row_vals)) >= 3:
            header_row = i
            break
    if header_row is None:
        header_row = 0  # fallback

    # 2) Re-read with the detected header
    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, header=header_row, dtype=str, engine="openpyxl")

    # Normalize headers and strip cells
    df.columns = [str(c).strip() for c in df.columns]
    df = df.applymap(lambda v: v.strip() if isinstance(v, str) else v)

    cols = list(df.columns)

    # 3) Find the invoice column
    # Preferred: second column by position (index 1), because user confirmed 2nd '#' is invoice no.
    inv_by_pos = None
    if len(cols) > 1:
        inv_by_pos = df.columns[1]

    # Named fallbacks (pandas often renames duplicate '#' to '#.1')
    name_candidates = ['#.1', '#2', 'Invoice', 'Invoice No', 'Invoice Number', 'Doc Num', 'DocNumber', '#']
    inv_by_name = next((c for c in cols if c in name_candidates or c.lower() in {
        '#.1', 'invoice', 'invoice no', 'invoice number', 'doc num', 'docnumber'
    }), None)

    invoice_col = inv_by_pos or inv_by_name
    if invoice_col is None:
        # last fallback: if there are at least 2 columns, force index 1
        if len(cols) >= 2:
            invoice_col = df.columns[1]
        else:
            raise ValueError(
                f"Could not find the invoice number column. Seen columns: {cols}"
            )

    # 4) Resolve the rest by relaxed name matching
    def _find(*names):
        lowers = [n.lower() for n in names]
        for c in cols:
            lc = c.lower()
            if lc in lowers:
                return c
        # try common variants
        variants = {
            "date": {"date", "posting date", "document date", "doc date"},
            "customer name": {"customer name", "bp name", "bp", "customer"},
            "sales employee": {"sales employee", "salesman", "sales emp"},
            "cancelled": {"cancelled", "canceled", "cancel", "cancellation"},
            "document total": {"document total", "doc total", "total", "amount"},
        }
        key = names[0].lower()
        for v in variants.get(key, set()):
            for c in cols:
                if c.lower() == v:
                    return c
        return None

    c_date  = _find("date")
    c_cust  = _find("customer name")
    c_sales = _find("sales employee")
    c_canc  = _find("cancelled")
    c_total = _find("document total")

    missing = [n for n, v in [
        ("Date", c_date), ("Customer Name", c_cust), ("Sales Employee", c_sales),
        ("Cancelled", c_canc), ("Document Total", c_total)
    ] if v is None]
    if missing:
        raise ValueError(f"Missing expected columns: {', '.join(missing)}. Seen: {cols}")

    # 5) Build cleaned frame
    out = pd.DataFrame({
        "invoice_number": df[invoice_col],
        "date_raw": df[c_date],
        "customer_name": df[c_cust],
        "salesman": df[c_sales].fillna(""),
        "cancelled_raw": df[c_canc],
        "document_total_raw": df[c_total],
    })

    # Keep only Cancelled == 'No' (case/space-insensitive)
    out = out[out["cancelled_raw"].astype(str).str.strip().str.lower() == "no"].copy()

    # Parse dates robustly: dd.mm.yy or dd.mm.yyyy or Excel serials already parsed
    parsed = pd.to_datetime(out["date_raw"], dayfirst=True, errors="coerce")
    # If parsing failed and values look like YYYY-MM-DD strings, try again without dayfirst
    mask_bad = parsed.isna() & out["date_raw"].astype(str).str.match(r"^\d{4}-\d{2}-\d{2}$")
    if mask_bad.any():
        parsed.loc[mask_bad] = pd.to_datetime(out.loc[mask_bad, "date_raw"], errors="coerce")
    out["date"] = parsed.dt.date

    # Amount
    out["document_total"] = out["document_total_raw"].map(_coerce_decimal)

    # Final trims
    out["invoice_number"] = out["invoice_number"].astype(str).str.strip()
    out["customer_name"]  = out["customer_name"].astype(str).str.strip()
    out["salesman"]       = out["salesman"].astype(str).str.strip()

    # Drop invalids
    out = out[(out["invoice_number"] != "") & out["date"].notna()]

    # Return only required columns
    return out[["invoice_number", "date", "customer_name", "salesman", "document_total"]]


# ---------- Views ----------
# orders/views.py
@transaction.atomic
def sap_invoices_upload(request):
    if request.method == "POST":
        form = SAPInvoiceUploadForm(request.POST, request.FILES)
        if form.is_valid():
            f = request.FILES["file"]
            try:
                df = _read_sap_dataframe(f)
            except Exception as e:
                messages.error(request, f"Upload failed: {e}")
                return redirect("sap_invoices_upload")

            batch = SAPInvoiceUploadBatch.objects.create(
                filename=getattr(f, "name", "sap_invoices.xlsx"),
                note=form.cleaned_data.get("note", ""),
                rows_ingested=0,
            )

            # --- metrics for transparency ---
            total_rows = len(df)
            inserted = 0
            updated = 0

            for row in df.to_dict(orient="records"):
                obj, created = SAPInvoice.objects.update_or_create(
                    invoice_number=row["invoice_number"],
                    defaults={
                        "date": row["date"],
                        "customer_name": row["customer_name"],
                        "salesman": row["salesman"],
                        "cancelled": False,  # only 'No' rows reach here
                        "document_total": row["document_total"],
                        "upload_batch": batch,
                    }
                )
                if created: inserted += 1
                else: updated += 1

            batch.rows_ingested = inserted
            batch.save(update_fields=["rows_ingested"])

            messages.success(
                request,
                f"Upload OK: total parsed {total_rows}, inserted {inserted}, updated {updated}."
            )
            # 👇 redirect to the LIST (or your SAP frequency page if you prefer)
            return redirect("customer_frequency_analysis_sap")
    else:
        form = SAPInvoiceUploadForm()
    return render(request, "sap_invoices/upload.html", {"form": form})

# views.py
from django.db import transaction
from django.contrib import messages
from django.shortcuts import redirect, render
from .utils import _read_sap_credit_dataframe
@transaction.atomic
def sap_credit_upload(request):
    if request.method == "POST":
        form = SAPInvoiceUploadForm(request.POST, request.FILES)
        if form.is_valid():
            f = request.FILES["file"]
            try:
                df = _read_sap_credit_dataframe(f)
            except Exception as e:
                messages.error(request, f"Credit upload failed: {e}")
                return redirect("sap_credit_upload")

            batch = SAPCreditNoteUploadBatch.objects.create(
                filename=getattr(f, "name", "sap_credit_notes.xlsx"),
                note=form.cleaned_data.get("note", ""),
                rows_ingested=0,
            )

            total_rows = len(df)
            inserted = 0
            updated = 0

            for row in df.to_dict(orient="records"):
                obj, created = SAPCreditNote.objects.update_or_create(
                    number=row["number"],
                    defaults={
                        "date": row["date"],
                        "customer_name": row["customer_name"],
                        "salesman": row["salesman"],
                        "document_total": row["document_total"],  # positive; subtracted in analysis
                        "upload_batch": batch,
                    }
                )
                if created: inserted += 1
                else: updated += 1

            batch.rows_ingested = inserted
            batch.save(update_fields=["rows_ingested"])

            messages.success(
                request,
                f"Credit upload OK: total parsed {total_rows}, inserted {inserted}, updated {updated}."
            )
            return redirect("customer_frequency_analysis_sap")
    else:
        form = SAPInvoiceUploadForm()
    return render(request, "sap_invoices/upload_credit.html", {"form": form, "page_title": "Upload SAP Credit Notes"})


 
def sap_invoices_list(request):
    """Very simple list page to verify uploads worked."""
    qs = SAPInvoice.objects.order_by("-date", "-created_at")[:500]
    return render(request, "sap_invoices/list.html", {"invoices": qs})
# --- put this near the top of views.py (module scope) ---
SALES_USER_MAP = {
    "muzain": ["B.MR.MUZAIN"],
    "dip": ["D.RETAIL CUST DIP"],
    "abubaqar": ["B. MR.RAFIQ ABU- PROJ","A.MR.RAFIQ ABU-TRD"],
    "rashid": ["A.MR.RASHID", "A.MR.RASHID CONT"],
    "parthiban": ["B.MR.PARTHIBAN"],
    "siyab": ["A.MR.SIYAB", "A.MR.SIYAB CONT"],
    "mr. nasheer": ["B.MR.NASHEER AHMAD"],
    "deira 2 store": ["R.DEIRA 2"],
    "rafiq": ["A.MR.RAFIQ"],
    "krishnan": ["I.KRISHNAN", "A.KRISHNAN"],
    "alabama": ["D. ALABAMA"],
    "anish": ["ANISH DIP"],
    "musharaf": ["A.MUSHARAF"],
    "ibrahim": ["A.IBRAHIM"],
    "adil": ["A.DIP ADIL"],
    "kadar": ["A.DIP KADAR"],
    "stephy": ["A.DIP STEFFY"],
    "muzammil": ["A.DIP MUZAMMIL"],
}

def _is_admin(user):
    uname = (getattr(user, "username", "") or "").lower().strip()
    return bool(getattr(user, "is_superuser", False) or getattr(user, "is_staff", False) or uname == "admin")


from datetime import datetime
from django.db.models import Count, Sum, Value, DecimalField, Q
from django.db.models.functions import TruncMonth, Coalesce
from django.shortcuts import render
from django.utils import timezone
from dateutil.relativedelta import relativedelta

PAGE_SIZE = 200  # tune as needed

def _parse_cursor(cursor: str | None):
    if not cursor:
        return None, None
    try:
        customer, salesman = cursor.split("||", 1)
        return customer, (None if salesman == "__NULL__" else salesman)
    except Exception:
        return None, None

def _make_cursor(customer_name: str, salesman: str | None):
    return f"{customer_name}||{salesman or '__NULL__'}"

# views.py
from datetime import datetime
from django.db.models import Count, Sum, Value, DecimalField, Q
from django.db.models.functions import TruncMonth, Coalesce
from django.shortcuts import render
from django.utils import timezone
from dateutil.relativedelta import relativedelta

PAGE_SIZE = 200  # keep your page size

def customer_frequency_analysis_sap(request):
    # Inputs
    start_month_str = request.GET.get('start')
    end_month_str   = request.GET.get('end')
    name_q          = (request.GET.get('q') or "").strip()
    after           = request.GET.get("after")
    salesman_filter = (request.GET.get('salesman') or "").strip()

    # Resolve user scope
    user = request.user
    is_admin = _is_admin(user)
    uname = (getattr(user, "username", "") or "").lower().strip()
    allowed_salesmen = SALES_USER_MAP.get(uname, [])

    # Date range (default: last 6 months inclusive)
    today = timezone.now().date().replace(day=1)
    if start_month_str and end_month_str:
        s_y, s_m = map(int, start_month_str.split("-"))
        e_y, e_m = map(int, end_month_str.split("-"))
        start_date = datetime(s_y, s_m, 1).date()
        end_month_first = datetime(e_y, e_m, 1).date()
        end_date = (end_month_first + relativedelta(months=1)) - relativedelta(days=1)
    else:
        start_date = (today - relativedelta(months=5))
        end_date = (today + relativedelta(months=1)) - relativedelta(days=1)

    # Base filtered queryset (Invoices)
    base = (SAPInvoice.objects
            .only("id", "date", "customer_name", "salesman", "document_total")
            .filter(date__range=[start_date, end_date]))

    # Enforce visibility
    if not is_admin:
        if allowed_salesmen:
            base = base.filter(salesman__in=allowed_salesmen)
        else:
            base = base.none()

    # Salesman param
    if salesman_filter:
        if is_admin:
            base = base.filter(salesman__iexact=salesman_filter)
        else:
            if salesman_filter in allowed_salesmen:
                base = base.filter(salesman__iexact=salesman_filter)

    if name_q:
        base = base.filter(customer_name__icontains=name_q)

    # ===== Credits with the same scope & filters (subtract by (customer, salesman)) =====
    credits = SAPCreditNote.objects.filter(date__range=[start_date, end_date])

    if not is_admin:
        if allowed_salesmen:
            credits = credits.filter(salesman__in=allowed_salesmen)
        else:
            credits = credits.none()

    if salesman_filter:
        if is_admin:
            credits = credits.filter(salesman__iexact=salesman_filter)
        else:
            if salesman_filter in allowed_salesmen:
                credits = credits.filter(salesman__iexact=salesman_filter)

    if name_q:
        credits = credits.filter(customer_name__icontains=name_q)

    # Build map safely (no 3-tuple dict() issue)
    credit_rows = (credits
                   .values("customer_name", "salesman")
                   .annotate(csum=Coalesce(Sum("document_total"),
                                           Value(0, output_field=DecimalField(max_digits=18, decimal_places=2)))))
    credit_pairs = { (r["customer_name"], r["salesman"]): float(r["csum"] or 0) for r in credit_rows }

    total_credits_all = float(
        credits.aggregate(tv=Coalesce(Sum("document_total"),
                                      Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))))["tv"] or 0
    )
    # ================================================================================

    # ----- GLOBAL STATS (independent of pagination) -----
    total_months_qs = (base
                       .annotate(m=TruncMonth("date"))
                       .values_list("m", flat=True)
                       .distinct())
    total_months = list(total_months_qs)
    total_months_count = len(total_months)

    total_value_all = (base
                       .aggregate(tv=Coalesce(
                           Sum("document_total"),
                           Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
                       ))["tv"] or 0)

    grouped_orders_all = (
        base.values("customer_name", "salesman")
            .annotate(orders=Count("id"))
    )
    one_time_count = sum(1 for r in grouped_orders_all if r["orders"] == 1)
    two_time_count = sum(1 for r in grouped_orders_all if r["orders"] == 2)

    customer_month_counts = (
        base.annotate(m=TruncMonth("date"))
            .values("customer_name")
            .annotate(mcnt=Count("m", distinct=True))
            .values("mcnt")
    )
    one_month_count = sum(1 for r in customer_month_counts if r["mcnt"] == 1)
    two_month_count = sum(1 for r in customer_month_counts if r["mcnt"] == 2)
    all_month_count = (
        sum(1 for r in customer_month_counts
            if total_months_count and r["mcnt"] == total_months_count)
        if total_months_count else 0
    )

    stats = {
        "one_month": one_month_count,
        "two_month": two_month_count,
        "all_month": all_month_count,
        "one_time": one_time_count,
        "two_time": two_time_count,
        # Net: invoices - credits (same scope)
        "total_value": float(total_value_all) - total_credits_all,
    }

    # ----- PAGE DATA (keyset pagination) -----
    grouped = (
        base.values("customer_name", "salesman")
            .annotate(
                orders=Count("id"),
                total_value=Coalesce(
                    Sum("document_total"),
                    Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
                )
            )
            .order_by("customer_name", "salesman")
    )

    if after:
        c_after, s_after = _parse_cursor(after)
        if c_after is not None:
            grouped = grouped.filter(
                Q(customer_name__gt=c_after) |
                (Q(customer_name=c_after) & Q(salesman__gt=s_after))
            )

    rows = list(grouped[:PAGE_SIZE + 1])
    has_next = len(rows) > PAGE_SIZE
    if has_next:
        last_included = rows[PAGE_SIZE - 1]
        next_cursor = _make_cursor(last_included["customer_name"], last_included["salesman"])
        rows = rows[:PAGE_SIZE]
    else:
        next_cursor = None

    # Months for only the current page pairs
    pairs = [(r["customer_name"], r["salesman"]) for r in rows]
    months_map = {}
    if pairs:
        q = Q()
        for cust, sman in pairs:
            q |= (Q(customer_name=cust) & Q(salesman=sman))
        month_rows = (
            base.filter(q)
                .annotate(m=TruncMonth("date"))
                .values("customer_name", "salesman", "m")
                .distinct()
        )
        for mr in month_rows:
            key = (mr["customer_name"], mr["salesman"])
            months_map.setdefault(key, []).append(mr["m"])

    # Build page results (subtract exact pair credit)
    results = []
    for r in rows:
        cust = r["customer_name"]
        sman = r["salesman"] or ""
        orders = r["orders"]
        inv_total = float(r["total_value"] or 0)

        csum = credit_pairs.get((cust, r["salesman"]), 0.0)
        net_total = inv_total - csum

        mlist = sorted(months_map.get((cust, r["salesman"]), []))
        results.append({
            "name": cust,
            "salesman": sman,
            "orders": orders,
            "months": [m.strftime("%b-%Y") for m in mlist],
            "order_class": "One-Time Customer" if orders == 1 else ("Two-Time Customer" if orders == 2 else f"{orders} Orders"),
            "month_class": (
                "One-Month Customer" if len(mlist) == 1 else
                "Two-Month Customer" if len(mlist) == 2 else
                ("All-Month Customer" if total_months_count and len(mlist) == total_months_count else f"{len(mlist)} Months")
            ),
            "total_value": round(net_total, 2),
        })

    # Salesmen dropdown
    if is_admin:
        salesmen = (SAPInvoice.objects
                    .exclude(salesman__isnull=True)
                    .exclude(salesman="")
                    .values_list("salesman", flat=True)
                    .distinct()
                    .order_by("salesman"))
    else:
        if allowed_salesmen:
            present = base.values_list("salesman", flat=True).distinct()
            present_set = set(present)
            salesmen = [s for s in allowed_salesmen if s in present_set]
        else:
            salesmen = []

    return render(request, "sap_invoices/customer_frequency_analysis_sap.html", {
        "results": results,
        "stats": stats,
        "salesmen": salesmen,
        "selected_salesman": salesman_filter if (is_admin or salesman_filter in (allowed_salesmen or [])) else "",
        "start": start_date.strftime("%Y-%m"),
        "end": end_date.strftime("%Y-%m"),
        "total_months": sorted([m.strftime("%b-%Y") for m in total_months]),
        "source": "SAP",
        "has_next": has_next,
        "next_cursor": next_cursor,
        "q": name_q,
    })



def customer_frequency_export_sap(request):
    # same filters as above
    resp = HttpResponse(content_type='text/csv')
    resp['Content-Disposition'] = 'attachment; filename="customer_frequency_sap.csv"'

    # reuse the query to avoid duplicate logic
    request.GET = request.GET.copy()  # ensure mutable for any tweaks
    # small inline compute (not factoring to keep this answer compact)
    start_month_str = request.GET.get('start')
    end_month_str = request.GET.get('end')
    today = timezone.now().date().replace(day=1)
    if start_month_str and end_month_str:
        s_y, s_m = map(int, start_month_str.split("-"))
        e_y, e_m = map(int, end_month_str.split("-"))
        start_date = datetime(s_y, s_m, 1).date()
        end_date = (datetime(e_y, e_m, 1).date() + relativedelta(months=1)) - relativedelta(days=1)
    else:
        start_date = (today - relativedelta(months=5))
        end_date = (today + relativedelta(months=1)) - relativedelta(days=1)

    salesman_filter = (request.GET.get('salesman') or "").strip()
    base = SAPInvoice.objects.filter(date__range=[start_date, end_date])
    if salesman_filter:
        base = base.filter(salesman__iexact=salesman_filter)

    agg = (base
           .values('customer_name', 'salesman')
           .annotate(
               orders=Count('id'),
               total_value = Coalesce(Sum('document_total'), Value(0, output_field=DecimalField()))
           ))

    months_map = defaultdict(set)
    for row in base.annotate(m=TruncMonth('date')).values('customer_name', 'm'):
        months_map[row['customer_name']].add(row['m'])

    import csv
    w = csv.writer(resp)
    w.writerow(['Customer','Salesman','Total Invoices','Month Classification','Months Bought','Total Value'])
    for r in agg:
        cust = r['customer_name']
        mset = months_map[cust]
        mcnt = len(mset)
        total_months = set(base.annotate(m=TruncMonth('date')).values_list('m', flat=True).distinct())
        if mcnt == 1: month_class = "One-Month Customer"
        elif mcnt == 2: month_class = "Two-Month Customer"
        elif len(total_months) and mcnt == len(total_months): month_class = "All-Month Customer"
        else: month_class = f"{mcnt} Months"
        w.writerow([cust, r['salesman'] or "", r['orders'],
                    month_class, "; ".join(sorted(m.strftime("%b-%Y") for m in mset)),
                    float(r['total_value'] or 0)])
    return resp
